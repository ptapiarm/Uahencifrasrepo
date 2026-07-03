"""
uah_admision.py  —  UAH en Cifras
"""

import io, base64
from pathlib import Path
from cryptography.fernet import Fernet, InvalidToken
import pandas as pd
import streamlit as st
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from sidebar_uah import render_sidebar


st.set_page_config(
    page_title="UAH · Indicadores Pregrado",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="collapsed",
)

 
# 1. set_page_config SIEMPRE primero, antes de cualquier otro comando.
st.set_page_config(page_title="Página de ejemplo — UAH", page_icon="🧩", layout="wide")
 
# 2. render_sidebar() justo después, para que el logo y el fondo negro
#    aparezcan también en esta página.
render_sidebar()
 
# ──────────────────────────────────────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&family=Barlow:wght@400;500;600&display=swap');

:root {
    --azul:        #1F3864;
    --azul-medio:  #2F5496;
    --azul-claro:  #BDD7EE;
    --azul-xclaro: #DEEAF1;
    --gris-fondo:  #F4F6F9;
    --gris-borde:  #C9D3DF;
    --gris-texto:  #555F6D;
    --acento:      #C00000;
}
html, body, [data-testid="stAppViewContainer"] {
    background: var(--gris-fondo);
    font-family: 'Barlow', sans-serif;
}

/* ── Header ── */
.uah-header {
    background: var(--azul); color: white;
    padding: 22px 32px 16px; border-radius: 12px; margin-bottom: 20px;
    display: flex; align-items: center; gap: 20px;
}
.uah-logo {
    font-family: 'Barlow Condensed', sans-serif; font-size: 3rem;
    font-weight: 700; line-height: 1;
    border-right: 3px solid rgba(255,255,255,.25); padding-right: 20px;
}
.uah-title { font-family:'Barlow Condensed',sans-serif; font-size:1.4rem;
              font-weight:600; text-transform:uppercase; letter-spacing:.5px;
              color:var(--azul-claro); }
.uah-desc  { font-size:.82rem; color:rgba(255,255,255,.6); margin-top:3px; }

/* ── Upload panel ── */
.upload-panel {
    background: white; border-radius: 10px; padding: 20px 24px;
    border: 1px solid var(--gris-borde); margin-bottom: 16px;
}
.upload-panel h4 {
    font-family:'Barlow Condensed',sans-serif; font-size:1.05rem;
    font-weight:700; letter-spacing:.3px;
    color:var(--azul); margin:0 0 12px 0;
}
.sheet-badge {
    background:var(--azul-xclaro); border-radius:7px; padding:11px 16px;
    border-left:4px solid var(--azul-medio); margin-bottom:16px;
    font-size:.84rem; color:var(--azul);
}

/* ── Tabs ── */
div[data-testid="stTabs"] button {
    font-family: 'Barlow Condensed', sans-serif !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    color: var(--azul) !important;
    padding: 8px 20px !important;
}
div[data-testid="stTabs"] button[aria-selected="true"] {
    color: var(--azul) !important;
    border-bottom: 3px solid var(--azul) !important;
}

/* ── Métricas ── */
.metric-strip { display:flex; gap:13px; margin-bottom:16px; flex-wrap:wrap; }
.metric-card  {
    background:white; border-radius:8px; padding:11px 17px;
    border-left:4px solid var(--azul-medio);
    box-shadow:0 1px 4px rgba(0,0,0,.07); min-width:130px;
}
.mc-label { font-size:.68rem; font-weight:600; text-transform:uppercase;
            letter-spacing:.8px; color:var(--gris-texto); margin-bottom:3px; }
.mc-value { font-family:'Barlow Condensed',sans-serif; font-size:1.85rem;
            font-weight:700; color:var(--azul); line-height:1; }
.metric-card.accent { border-left-color:var(--acento); }
.metric-card.accent .mc-value { color:var(--acento); }

/* ── Tabla genérica ── */
.table-wrapper { background:white; border-radius:10px;
                 box-shadow:0 2px 8px rgba(0,0,0,.08); overflow:hidden;
                 margin-bottom: 16px; }
.table-title-bar {
    background:var(--azul); color:white; padding:12px 20px;
    font-family:'Barlow Condensed',sans-serif; font-size:.98rem;
    font-weight:600; letter-spacing:.5px; text-transform:uppercase;
}
.table-scroll { overflow-x:auto; max-height:66vh; overflow-y:auto; }
.table-scroll::-webkit-scrollbar { height:6px; width:6px; }
.table-scroll::-webkit-scrollbar-thumb { background:var(--azul-claro); border-radius:3px; }

/* ── Tabla 1 ── */
table.t1 { width:100%; border-collapse:collapse;
            font-size:.8rem; font-family:'Barlow',sans-serif; }
table.t1 thead th {
    background:var(--azul); color:white;
    font-family:'Barlow Condensed',sans-serif; font-weight:700; font-size:.92rem;
    padding:9px 7px; text-align:center; border:1px solid rgba(255,255,255,.15);
    white-space:nowrap; position:sticky; top:0; z-index:2;
}
table.t1 thead th.c-fac     { text-align:left; min-width:120px; }
table.t1 thead th.c-siga    { min-width:100px; }
table.t1 thead th.c-carrera { text-align:left; min-width:260px; }
table.t1 tbody tr { border-bottom:1px solid #EEF1F5; }
table.t1 tbody tr:hover td  { background:var(--azul-xclaro) !important; }
table.t1 tbody td { padding:7px 7px; border-right:1px solid #EEF1F5;
                    text-align:center; color:#2C3A50; }
table.t1 tbody td.c-fac {
    background:var(--azul-xclaro); font-weight:600; color:var(--azul);
    text-align:left; padding-left:12px; font-size:.77rem;
    text-transform:uppercase; letter-spacing:.4px;
}
table.t1 tbody td.c-siga    { background:#F8FAFC; color:var(--azul-medio);
                               font-weight:600; font-size:.77rem; }
table.t1 tbody td.c-carrera { text-align:left; padding-left:11px;
                               font-weight:500; color:#1E2D45; background:#FAFBFD; }
table.t1 tbody td.c-num     { font-family:'Barlow Condensed',sans-serif;
                               font-size:.88rem; font-weight:600; }
table.t1 tbody td.c-empty   { color:#CBD5E1; font-size:.68rem; }
table.t1 thead th.c-diff    { background:#1a3a6b; min-width:90px; }
table.t1 thead th.c-var     { background:#1a3a6b; min-width:80px; }
table.t1 tbody tr.tr-tot td { background:var(--azul-medio) !important;
                               color:white !important; font-weight:700;
                               font-family:'Barlow Condensed',sans-serif; }
table.t1 tbody tr.tr-tot td.c-fac,
table.t1 tbody tr.tr-tot td.c-carrera { background:var(--azul) !important; }

/* ── Tabla 2 ── */
table.t2 { width:100%; border-collapse:collapse;
           font-size:.79rem; font-family:'Barlow',sans-serif; }
table.t2 thead tr.rh1 th {
    background:var(--azul); color:white;
    font-family:'Barlow Condensed',sans-serif; font-weight:700; font-size:.9rem;
    padding:8px 6px; text-align:center; border:1px solid rgba(255,255,255,.15);
    position:sticky; top:0; z-index:3; white-space:nowrap;
    text-transform:none !important; font-variant:normal !important;
}
table.t2 thead tr.rh1 th.th-left { text-align:left; }
table.t2 thead tr.rh1 th.th-yr   { background:var(--azul-medio);
                                    border-left:2px solid rgba(255,255,255,.35); }
table.t2 thead tr.rh2 th {
    background:#2A4070; color:#D8E8F5;
    font-family:'Barlow Condensed',sans-serif; font-weight:600; font-size:.85rem;
    padding:5px 5px; text-align:center; border:1px solid rgba(255,255,255,.1);
    position:sticky; top:33px; z-index:2; white-space:nowrap;
    text-transform:none !important; font-variant:normal !important;
}
table.t2 thead tr.rh2 th.sh-muj  { color:#FFB3D1; }
table.t2 thead tr.rh2 th.sh-hom  { color:#A3C8FF; }
table.t2 thead tr.rh2 th.sh-pct  { color:#C8D8E8; font-style:italic; font-size:.78rem; }
table.t2 thead tr.rh2 th.sh-pace { color:#C8D8E8; font-size:.85rem; font-family:monospace; font-style:normal; font-weight:600; }
table.t2-via thead tr.rh2 th { font-style:normal !important; }
table.t2 thead tr.rh2 th.sh-tot  { color:white; font-weight:700; background:#1F3864; }
table.t2 thead tr.rh2 th.sh-fix  { background:#1a3055; }
table.t2 tbody tr { border-bottom:1px solid #EEF1F5; }
table.t2 tbody tr:hover td { background:var(--azul-xclaro) !important; }
table.t2 tbody td { padding:6px 6px; border-right:1px solid #EEF1F5;
                    text-align:center; color:#2C3A50; }
table.t2 tbody td.c-fac {
    background:var(--azul-xclaro); font-weight:600; color:var(--azul);
    text-align:left; padding-left:12px; font-size:.76rem;
    text-transform:uppercase; letter-spacing:.3px;
}
table.t2 tbody td.c-jorn { background:#F0F4F8; color:#4A6080;
                            font-size:.74rem; font-weight:500; }
table.t2 tbody td.c-siga { background:#F8FAFC; color:var(--azul-medio);
                            font-weight:600; font-size:.76rem; }
table.t2 tbody td.c-carr { text-align:left; padding-left:10px;
                            font-weight:500; color:#1E2D45; background:#FAFBFD;
                            max-width:210px; overflow:hidden;
                            text-overflow:ellipsis; white-space:nowrap; }
table.t2 tbody td.c-muj  { font-family:'Barlow Condensed',sans-serif;
                            font-size:.88rem; font-weight:700; color:#8B2252; }
table.t2 tbody td.c-hom  { font-family:'Barlow Condensed',sans-serif;
                            font-size:.88rem; font-weight:700; color:#1A4B8C; }
table.t2 tbody td.c-pct  { font-size:.76rem; color:#556677; }
table.t2 tbody td.c-tot  { font-family:'Barlow Condensed',sans-serif;
                            font-size:.9rem; font-weight:700; color:var(--azul);
                            background:#F0F5FA;
                            border-left:2px solid var(--azul-claro) !important; }
table.t2 tbody td.c-sep  { border-left:2px solid var(--azul-claro) !important; }
table.t2 tbody td.c-empty{ color:#CBD5E1; font-size:.67rem; }
table.t2 tbody tr.tr-tot td { background:var(--azul-medio) !important;
                               color:white !important; font-weight:700;
                               font-family:'Barlow Condensed',sans-serif; }
table.t2 tbody tr.tr-tot td.c-fac,
table.t2 tbody tr.tr-tot td.c-carr { background:var(--azul) !important; }
table.t2 tbody tr.tr-tot td.c-pct  { color:rgba(255,255,255,.8) !important; }
table.t2 tbody tr.tr-tot td.c-tot  { background:var(--azul) !important;
                                      color:white !important; }

/* ── Botones de descarga ── */
.dl-btn {
    display: inline-block; background: var(--azul); color: white !important;
    text-decoration: none !important; font-family: 'Barlow', sans-serif;
    font-weight: 600; font-size: .84rem; padding: 8px 18px;
    border-radius: 6px; margin-right: 10px; transition: background .2s;
}
.dl-btn:hover { background: var(--azul-medio); color: white !important; }
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# CARGA DE ARCHIVO INSTITUCIONAL CIFRADO (data/*.enc)
# ──────────────────────────────────────────────────────────────────────────────
# Los archivos .enc dentro de la carpeta data/ son versiones cifradas de los
# Excel institucionales (generadas con cifrar_excels.py). No son legibles sin
# la clave, así que es seguro tenerlos en un repositorio de GitHub aunque sea
# público. La clave SOLO vive en .streamlit/secrets.toml (nunca en el repo):
#
# [encryption]
# key = "el-contenido-de-clave.key"


def _buscar_archivo_enc(nombre_archivo: str):
    """
    Busca nombre_archivo dentro de una carpeta 'data' en varios lugares
    posibles, porque en una app multipágina este script vive en pages/,
    no en la raíz del repo:
      - pages/data/archivo.enc      (al lado del propio script)
      - data/archivo.enc            (un nivel arriba, raíz del repo)
      - <cwd>/data/archivo.enc      (directorio de trabajo de Streamlit)
    Devuelve el primer Path que exista, o None si no lo encuentra en ninguno.
    """
    aqui = Path(__file__).resolve().parent
    candidatos = [
        aqui / "data" / nombre_archivo,
        aqui.parent / "data" / nombre_archivo,
        Path.cwd() / "data" / nombre_archivo,
    ]
    for c in candidatos:
        if c.exists():
            return c
    return None


@st.cache_data(show_spinner=False)
def _descifrar_archivo(ruta_enc: str) -> bytes:
    p = Path(ruta_enc)
    if not p.exists():
        raise FileNotFoundError(f"No existe el archivo cifrado: {p}")
    try:
        clave = st.secrets["encryption"]["key"]
    except (KeyError, FileNotFoundError):
        raise RuntimeError(
            "Falta configurar [encryption] key en .streamlit/secrets.toml."
        )
    try:
        return Fernet(clave.encode() if isinstance(clave, str) else clave).decrypt(p.read_bytes())
    except InvalidToken:
        raise RuntimeError("La clave configurada no corresponde a este archivo cifrado.")


def selector_origen_archivo(nombre: str, ejemplo: str, key_prefix: str, archivo_enc: str = None):
    """
    Muestra el panel para cargar el archivo Excel de {nombre}: ya sea subiéndolo
    manualmente, o (si existe data/{archivo_enc}) descifrando la versión
    institucional incluida en la app. Devuelve (file_bytes, hoja_sel).
    """
    st.markdown(
        f'<div class="upload-panel"><h4>📂 Archivo de {nombre}</h4>'
        f'<p style="font-size:.8rem;color:#6B7A8D;margin:-6px 0 10px 0">📎 Ejemplo: <code>{ejemplo}</code></p>',
        unsafe_allow_html=True,
    )

    ruta_enc = _buscar_archivo_enc(archivo_enc) if archivo_enc else None
    hay_institucional = ruta_enc is not None

    if hay_institucional:
        origen = st.radio(
            "¿Cómo quieres cargar el archivo?",
            ["🔒  Usar archivo institucional (incluido en la app)", "💻  Subir desde mi computador"],
            horizontal=True, key=f"{key_prefix}_origen",
        )
    else:
        origen = "💻  Subir desde mi computador"

    col_up, col_sh = st.columns([3, 2])
    file_bytes = None
    with col_up:
        if origen.startswith("🔒"):
            try:
                file_bytes = _descifrar_archivo(str(ruta_enc))
                st.success("Archivo institucional cargado ✅")
            except Exception as e:
                st.error(f"No se pudo cargar el archivo institucional.\n\n`{e}`")
        else:
            uploaded = st.file_uploader(
                f"Sube el Excel de {nombre}", type=["xlsx"],
                label_visibility="visible", key=f"{key_prefix}_uploader",
            )
            if uploaded:
                file_bytes = uploaded.read()

    with col_sh:
        if file_bytes:
            hojas = get_sheet_names(file_bytes)
            hoja_sel = st.selectbox(
                "Seleccionar hoja de trabajo", hojas,
                index=len(hojas) - 1, key=f"{key_prefix}_hoja",
            )
        else:
            hoja_sel = None
            st.selectbox(
                "Seleccionar hoja de trabajo",
                ["— sube un archivo primero —"], disabled=True, key=f"{key_prefix}_hoja_dis",
            )
    st.markdown('</div>', unsafe_allow_html=True)
    return file_bytes, hoja_sel


# ──────────────────────────────────────────────────────────────────────────────
# FUNCIONES DE DATOS
# ──────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def get_sheet_names(file_bytes: bytes) -> list:
    return pd.ExcelFile(io.BytesIO(file_bytes)).sheet_names


@st.cache_data(show_spinner=False)
def procesar_admision(file_bytes: bytes, hoja: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=4)
    df = df[df["AÑO INGRESO"].between(2000, 2100)].copy()
    df["AÑO INGRESO"] = df["AÑO INGRESO"].astype(int)
    grp = (df.groupby(["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL", "AÑO INGRESO"])
           ["RUT"].count().reset_index(name="N"))
    pivot = grp.pivot_table(
        index=["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"],
        columns="AÑO INGRESO", values="N", aggfunc="sum"
    ).reset_index()
    pivot.columns.name = None
    pivot["CODIGO SIGA"] = pivot["CODIGO SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    return pivot.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def procesar_admision_regular(file_bytes: bytes, hoja: str) -> pd.DataFrame:
    """Igual que procesar_admision pero filtra solo filas con indicador '1.OK'."""
    IND_COL = "INDICADOR DEFINITIVO INDICADORES ADMISIÓN 30 DE ABRIL"
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=4)
    df = df[df["AÑO INGRESO"].between(2000, 2100)].copy()
    df["AÑO INGRESO"] = df["AÑO INGRESO"].astype(int)
    df = df[df[IND_COL].astype(str).str.startswith("1.OK")].copy()
    grp = (df.groupby(["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL", "AÑO INGRESO"])
             ["RUT"].count().reset_index(name="N"))
    pivot = grp.pivot_table(
        index=["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"],
        columns="AÑO INGRESO", values="N", aggfunc="sum"
    ).reset_index()
    pivot.columns.name = None
    pivot["CODIGO SIGA"] = pivot["CODIGO SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    return pivot.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def procesar_via_ingreso(file_bytes: bytes, hoja: str) -> pd.DataFrame:
    VIA_COL  = "VIA DE INGRESO (DEMRE-PACE-ESPECIAL-COMPL-CONDUC)"
    VIA_MAP  = {
        "1. DEMRE":              "Adm. Ordinaria DEMRE",
        "3. ADMISIÓN ESPECIAL":  "Adm. Especial (*)",
        "4. COMPLEMENTARIA":     "Complementaria",
        "5. CONDUCENCIA":        "Conducencia",
        "2. PACE":               "PACE",
        "DEMRE":                 "Adm. Ordinaria DEMRE",
        "ESPECIAL":              "Adm. Especial (*)",
        "COMPLEMENTARIA":        "Complementaria",
        "CONDUCENCIA":           "Conducencia",
        "PACE":                  "PACE",
    }
    VIA_ORDER = ["Adm. Ordinaria DEMRE", "Adm. Especial (*)", "Complementaria", "Conducencia", "PACE"]

    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=4,
                       usecols=["AÑO INGRESO", "CODIGO SIGA", "FACULTAD NOMPROPIO",
                                "NOMBRE CARRERA FINAL", "RUT", VIA_COL])
    df = df[pd.to_numeric(df["AÑO INGRESO"], errors="coerce").between(2000, 2100)].copy()
    df["AÑO INGRESO"] = df["AÑO INGRESO"].astype(int)
    df["CODIGO SIGA"] = pd.to_numeric(df["CODIGO SIGA"], errors="coerce")
    def map_via(v):
        v = str(v).strip()
        if "DEMRE" in v:             return "Adm. Ordinaria DEMRE"
        if "ESPECIAL" in v:          return "Adm. Especial (*)"
        if "COMPLEMENTARIA" in v:    return "Complementaria"
        if "CONDUCENCIA" in v:       return "Conducencia"
        if "PACE" in v : return "PACE"
        return "Otra"

    df["VIA"] = df[VIA_COL].apply(map_via)

    # Meta canónica por código
    meta = (df.sort_values("AÑO INGRESO")
              .drop_duplicates(subset="CODIGO SIGA", keep="last")
              [["CODIGO SIGA", "FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]])

    # Contar por código/año/vía
    grp = (df.groupby(["CODIGO SIGA", "AÑO INGRESO", "VIA"], observed=True)
             ["RUT"].count().reset_index(name="N"))
    tot = (df.groupby(["CODIGO SIGA", "AÑO INGRESO"])
             ["RUT"].count().reset_index(name="Total"))

    grp = grp.merge(tot, on=["CODIGO SIGA", "AÑO INGRESO"])
    grp = grp.merge(meta, on="CODIGO SIGA")
    grp["CODIGO SIGA"] = grp["CODIGO SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    return grp.sort_values(
        ["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL", "AÑO INGRESO", "VIA"]
    ).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def procesar_sexo(file_bytes: bytes, hoja: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=4)
    df = df[df["AÑO INGRESO"].between(2000, 2100)].copy()
    df["AÑO INGRESO"] = df["AÑO INGRESO"].astype(int)
    df["SEXO FINAL"]  = df["SEXO FINAL"].astype(str).str.strip()

    def fmt_jornada(v):
        if not isinstance(v, str):
            return None
        v = v.strip()
        if "distancia" in v.lower():
            return "ONLINE"
        return v.upper()

    df["JORNADA"] = df["JORNADA FICHA"].apply(fmt_jornada).fillna("")
    jornada_label = (
        df[df["JORNADA"] != ""]
        .groupby("CODIGO SIGA")["JORNADA"]
        .first()
        .to_dict()
    )

    df_s = df[df["SEXO FINAL"].isin(["M", "H"])].copy()
    grp  = (df_s.groupby([
        "FACULTAD NOMPROPIO", "CODIGO SIGA",
        "NOMBRE CARRERA FINAL", "AÑO INGRESO", "SEXO FINAL"
    ])["RUT"].count().reset_index(name="N"))
    tot  = (df_s.groupby(["CODIGO SIGA", "AÑO INGRESO"])
            ["RUT"].count().reset_index(name="Total"))
    grp  = grp.merge(tot, on=["CODIGO SIGA", "AÑO INGRESO"])
    grp["PCT"] = (grp["N"] / grp["Total"]).round(4)
    grp["CODIGO SIGA"] = grp["CODIGO SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    grp["JORNADA"] = grp["CODIGO SIGA"].map(jornada_label).fillna("")
    return grp.sort_values(
        ["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL", "AÑO INGRESO", "SEXO FINAL"]
    ).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def contar_matriculados_por_anio(file_bytes: bytes, hoja: str) -> dict:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=2)
    df = df[pd.to_numeric(df["AÑO MATRÍCULA"], errors="coerce").between(2000, 2100)].copy()
    df["AÑO MATRÍCULA"] = df["AÑO MATRÍCULA"].astype(int)
    return (df.groupby("AÑO MATRÍCULA")["RUT SIGA BUSQUEDA DE DATOS"]
              .nunique().to_dict())


@st.cache_data(show_spinner=False)
def procesar_matricula(file_bytes: bytes, hoja: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=2)
    df = df[pd.to_numeric(df["AÑO MATRÍCULA"], errors="coerce").between(2000, 2100)].copy()
    df["AÑO MATRÍCULA"]  = df["AÑO MATRÍCULA"].astype(int)
    df["COD CARREA SIGA"] = pd.to_numeric(df["COD CARREA SIGA"], errors="coerce")

    # Meta canónica: una sola fila por código, tomando los valores del año más reciente
    meta = (df.sort_values("AÑO MATRÍCULA")
              .drop_duplicates(subset="COD CARREA SIGA", keep="last")
              [["COD CARREA SIGA", "FACULTAD", "NOMBRE CARRERA INFORMES"]]
              .reset_index(drop=True))

    # Contar RUTs únicos por código de carrera y año
    grp = (df.groupby(["COD CARREA SIGA", "AÑO MATRÍCULA"])
             ["RUT SIGA BUSQUEDA DE DATOS"].nunique()
             .reset_index(name="N"))

    # Pivot solo por código de carrera (una fila por código, sin riesgo de duplicados)
    pivot = grp.pivot_table(
        index="COD CARREA SIGA",
        columns="AÑO MATRÍCULA",
        values="N",
        aggfunc="sum"
    ).reset_index()
    pivot.columns.name = None

    # Unir meta al final
    pivot = pivot.merge(meta, on="COD CARREA SIGA")
    pivot["COD CARREA SIGA"] = pivot["COD CARREA SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    pivot = pivot.rename(columns={
        "FACULTAD": "FACULTAD NOMPROPIO",
        "COD CARREA SIGA": "CODIGO SIGA",
        "NOMBRE CARRERA INFORMES": "NOMBRE CARRERA FINAL",
    })
    return pivot.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def procesar_matricula_sexo(file_bytes: bytes, hoja: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=2)
    df = df[pd.to_numeric(df["AÑO MATRÍCULA"], errors="coerce").between(2000, 2100)].copy()
    df["AÑO MATRÍCULA"]   = df["AÑO MATRÍCULA"].astype(int)
    df["COD CARREA SIGA"]  = pd.to_numeric(df["COD CARREA SIGA"], errors="coerce")
    df["SEXO FINAL"]       = df["Sexo final"].astype(str).str.strip().str.upper()

    meta = (df.sort_values("AÑO MATRÍCULA")
              .drop_duplicates(subset="COD CARREA SIGA", keep="last")
              [["COD CARREA SIGA", "FACULTAD", "NOMBRE CARRERA INFORMES"]]
              .reset_index(drop=True))

    df_s = df[df["SEXO FINAL"].isin(["M", "H"])].copy()
    grp  = (df_s.groupby(["COD CARREA SIGA", "AÑO MATRÍCULA", "SEXO FINAL"])
               ["RUT SIGA BUSQUEDA DE DATOS"].nunique().reset_index(name="N"))
    tot  = (df_s.groupby(["COD CARREA SIGA", "AÑO MATRÍCULA"])
               ["RUT SIGA BUSQUEDA DE DATOS"].nunique().reset_index(name="Total"))
    grp  = grp.merge(tot, on=["COD CARREA SIGA", "AÑO MATRÍCULA"])
    grp  = grp.merge(meta, on="COD CARREA SIGA")
    grp["PCT"] = (grp["N"] / grp["Total"]).round(4)
    grp["COD CARREA SIGA"] = grp["COD CARREA SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    grp["JORNADA"] = ""
    grp = grp.rename(columns={
        "FACULTAD": "FACULTAD NOMPROPIO",
        "COD CARREA SIGA": "CODIGO SIGA",
        "NOMBRE CARRERA INFORMES": "NOMBRE CARRERA FINAL",
        "AÑO MATRÍCULA": "AÑO INGRESO",
    })
    return grp.sort_values(
        ["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL", "AÑO INGRESO", "SEXO FINAL"]
    ).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def procesar_matricula_nueva(file_bytes: bytes, hoja: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja, header=2)
    df = df[pd.to_numeric(df["AÑO MATRÍCULA"], errors="coerce").between(2000, 2100)].copy()
    df["AÑO MATRÍCULA"]   = df["AÑO MATRÍCULA"].astype(int)
    df["AÑO DE INGRESO"]  = pd.to_numeric(df["AÑO DE INGRESO"], errors="coerce")
    df["COD CARREA SIGA"] = pd.to_numeric(df["COD CARREA SIGA"], errors="coerce")

    # Clasificar: NUEVO = año ingreso == año matrícula, SUPERIOR = año ingreso < año matrícula
    df["TIPO"] = df.apply(
        lambda r: "NUEVO" if r["AÑO DE INGRESO"] == r["AÑO MATRÍCULA"] else "SUPERIOR", axis=1
    )

    meta = (df.sort_values("AÑO MATRÍCULA")
              .drop_duplicates(subset="COD CARREA SIGA", keep="last")
              [["COD CARREA SIGA", "FACULTAD", "NOMBRE CARRERA INFORMES"]]
              .reset_index(drop=True))

    grp = (df.groupby(["COD CARREA SIGA", "AÑO MATRÍCULA", "TIPO"])
             ["RUT SIGA BUSQUEDA DE DATOS"].nunique()
             .reset_index(name="N"))
    tot = (df.groupby(["COD CARREA SIGA", "AÑO MATRÍCULA"])
             ["RUT SIGA BUSQUEDA DE DATOS"].nunique()
             .reset_index(name="Total"))
    grp = grp.merge(tot, on=["COD CARREA SIGA", "AÑO MATRÍCULA"])
    grp["PCT"] = (grp["N"] / grp["Total"]).round(4)
    grp = grp.merge(meta, on="COD CARREA SIGA")
    grp["COD CARREA SIGA"] = grp["COD CARREA SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    grp = grp.rename(columns={
        "FACULTAD": "FACULTAD NOMPROPIO",
        "COD CARREA SIGA": "CODIGO SIGA",
        "NOMBRE CARRERA INFORMES": "NOMBRE CARRERA FINAL",
        "AÑO MATRÍCULA": "AÑO",
    })
    return grp.sort_values(
        ["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL", "AÑO", "TIPO"]
    ).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def procesar_retencion(file_bytes_adm: bytes, hoja_adm: str,
                       file_bytes_mat: bytes, hoja_mat: str) -> pd.DataFrame:
    # ── Admisión: cohortes válidas ──
    df_adm = pd.read_excel(io.BytesIO(file_bytes_adm), sheet_name=hoja_adm, header=4)
    df_adm = df_adm[pd.to_numeric(df_adm["AÑO INGRESO"], errors="coerce").between(2000, 2100)].copy()
    df_adm["AÑO INGRESO"]   = df_adm["AÑO INGRESO"].astype(int)
    df_adm["CODIGO SIGA"]   = pd.to_numeric(df_adm["CODIGO SIGA"], errors="coerce")
    df_adm["RUT"]           = df_adm["RUT"].astype(str).str.strip()
    ind_col = "INDICADOR DEFINITIVO INDICADORES ADMISIÓN 30 DE ABRIL"
    df_adm  = df_adm[df_adm[ind_col].astype(str).str.startswith("1.OK")].copy()

    # Meta canónica de carreras desde admisión
    meta = (df_adm.sort_values("AÑO INGRESO")
                  .drop_duplicates(subset="CODIGO SIGA", keep="last")
                  [["CODIGO SIGA", "FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]]
                  .reset_index(drop=True))

    # Tamaño de cohorte por carrera y año
    cohortes = (df_adm.groupby(["CODIGO SIGA", "AÑO INGRESO"])
                      ["RUT"].nunique().reset_index(name="N_COHORTE")
                      .rename(columns={"AÑO INGRESO": "AÑO COHORTE"}))

    # ── Matrícula: presencia año siguiente ──
    df_mat = pd.read_excel(io.BytesIO(file_bytes_mat), sheet_name=hoja_mat, header=2)
    df_mat = df_mat[pd.to_numeric(df_mat["AÑO MATRÍCULA"], errors="coerce").between(2000, 2100)].copy()
    df_mat["AÑO MATRÍCULA"]   = df_mat["AÑO MATRÍCULA"].astype(int)
    df_mat["AÑO DE INGRESO"]  = pd.to_numeric(df_mat["AÑO DE INGRESO"], errors="coerce")
    df_mat["COD CARREA SIGA"] = pd.to_numeric(df_mat["COD CARREA SIGA"], errors="coerce")
    df_mat["RUT"]             = df_mat["RUT SIGA BUSQUEDA DE DATOS"].astype(str).str.strip()

    # Estudiantes en matrícula con AÑO DE INGRESO == cohorte y AÑO MATRÍCULA == cohorte+1
    df_mat_ret = df_mat.drop_duplicates(subset=["RUT", "COD CARREA SIGA", "AÑO MATRÍCULA"])

    # Para cada cohorte X, buscar RUTs en matrícula con AÑO MATRÍCULA=X+1 y AÑO DE INGRESO=X
    # Cruzar admisión (cohorte X, RUT, CODIGO SIGA) con matrícula (AÑO MATRÍCULA=X+1)
    df_adm_ruts = (df_adm[["RUT", "CODIGO SIGA", "AÑO INGRESO"]]
                   .drop_duplicates()
                   .rename(columns={"AÑO INGRESO": "AÑO COHORTE"}))
    df_adm_ruts["AÑO_SIG"] = df_adm_ruts["AÑO COHORTE"] + 1

    mat_sig = (df_mat_ret[df_mat_ret["AÑO DE INGRESO"].notna()]
               .copy())
    mat_sig = mat_sig.rename(columns={
        "COD CARREA SIGA": "CODIGO SIGA",
        "AÑO MATRÍCULA": "AÑO_SIG",
        "AÑO DE INGRESO": "AÑO COHORTE",
    })[["RUT", "CODIGO SIGA", "AÑO_SIG", "AÑO COHORTE"]]
    mat_sig["AÑO COHORTE"] = mat_sig["AÑO COHORTE"].astype(int)

    ret = df_adm_ruts.merge(
        mat_sig[["RUT", "CODIGO SIGA", "AÑO_SIG", "AÑO COHORTE"]],
        on=["RUT", "CODIGO SIGA", "AÑO_SIG", "AÑO COHORTE"],
        how="left",
        indicator=True
    )
    ret["RETENIDO"] = (ret["_merge"] == "both").astype(int)

    ret_coh = (ret[ret["RETENIDO"] == 1]
               .groupby(["CODIGO SIGA", "AÑO COHORTE"])
               ["RUT"].nunique().reset_index(name="N_RETENIDOS"))

    res = cohortes.merge(ret_coh, on=["CODIGO SIGA", "AÑO COHORTE"], how="left")
    res["N_RETENIDOS"] = res["N_RETENIDOS"].fillna(0).astype(int)
    res["PCT"] = (res["N_RETENIDOS"] / res["N_COHORTE"]).round(4)
    res = res.merge(meta, on="CODIGO SIGA")
    res["CODIGO SIGA"] = res["CODIGO SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    return res.sort_values(
        ["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL", "AÑO COHORTE"]
    ).reset_index(drop=True)


@st.cache_data(show_spinner=False)
def procesar_seguimiento_cohorte(file_bytes_adm: bytes, hoja_adm: str,
                                  file_bytes_mat: bytes, hoja_mat: str) -> pd.DataFrame:
    """
    Para cada cohorte X y carrera, cuenta cuántos RUTs de esa cohorte
    aparecen en matrícula en los años X, X+1, X+2, ... (seguimiento longitudinal).
    Retorna un DataFrame largo con columnas:
        CODIGO SIGA, FACULTAD NOMPROPIO, NOMBRE CARRERA FINAL,
        AÑO COHORTE, N_COHORTE, AÑO MATRICULA, ANIOS_DESDE_INGRESO, N_PRESENTES
    """
    # ── Admisión: cohortes válidas (solo indicador OK) ──
    df_adm = pd.read_excel(io.BytesIO(file_bytes_adm), sheet_name=hoja_adm, header=4)
    df_adm = df_adm[pd.to_numeric(df_adm["AÑO INGRESO"], errors="coerce").between(2000, 2100)].copy()
    df_adm["AÑO INGRESO"] = df_adm["AÑO INGRESO"].astype(int)
    df_adm["CODIGO SIGA"] = pd.to_numeric(df_adm["CODIGO SIGA"], errors="coerce")
    df_adm["RUT"]         = df_adm["RUT"].astype(str).str.strip()
    ind_col = "INDICADOR DEFINITIVO INDICADORES ADMISIÓN 30 DE ABRIL"
    df_adm  = df_adm[df_adm[ind_col].astype(str).str.startswith("1.OK")].copy()

    # Meta canónica de carreras
    meta = (df_adm.sort_values("AÑO INGRESO")
                  .drop_duplicates(subset="CODIGO SIGA", keep="last")
                  [["CODIGO SIGA", "FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]]
                  .reset_index(drop=True))

    # Tamaño de cohorte por carrera y año
    cohortes = (df_adm.groupby(["CODIGO SIGA", "AÑO INGRESO"])
                      ["RUT"].nunique().reset_index(name="N_COHORTE")
                      .rename(columns={"AÑO INGRESO": "AÑO COHORTE"}))

    # RUTs por cohorte
    adm_ruts = (df_adm[["RUT", "CODIGO SIGA", "AÑO INGRESO"]]
                .drop_duplicates()
                .rename(columns={"AÑO INGRESO": "AÑO COHORTE"}))

    # ── Matrícula: todos los años ──
    df_mat = pd.read_excel(io.BytesIO(file_bytes_mat), sheet_name=hoja_mat, header=2)
    df_mat = df_mat[pd.to_numeric(df_mat["AÑO MATRÍCULA"], errors="coerce").between(2000, 2100)].copy()
    df_mat["AÑO MATRÍCULA"]   = df_mat["AÑO MATRÍCULA"].astype(int)
    df_mat["AÑO DE INGRESO"]  = pd.to_numeric(df_mat["AÑO DE INGRESO"], errors="coerce")
    df_mat["COD CARREA SIGA"] = pd.to_numeric(df_mat["COD CARREA SIGA"], errors="coerce")
    df_mat["RUT"]             = df_mat["RUT SIGA BUSQUEDA DE DATOS"].astype(str).str.strip()
    df_mat = df_mat.drop_duplicates(subset=["RUT", "COD CARREA SIGA", "AÑO MATRÍCULA"])

    # Renombrar para el merge
    mat_long = df_mat.rename(columns={
        "COD CARREA SIGA": "CODIGO SIGA",
        "AÑO MATRÍCULA":  "AÑO MATRICULA",
        "AÑO DE INGRESO": "AÑO COHORTE_MAT",
    })[["RUT", "CODIGO SIGA", "AÑO MATRICULA", "AÑO COHORTE_MAT"]]

    # Para cada cohorte X, contar los RUTs de adm presentes en cada año de matrícula
    # Cruzar: adm_ruts (RUT, SIGA, AÑO COHORTE) x mat_long (RUT, SIGA, AÑO MATRICULA)
    merged = adm_ruts.merge(
        mat_long[["RUT", "CODIGO SIGA", "AÑO MATRICULA"]],
        on=["RUT", "CODIGO SIGA"],
        how="inner"
    )

    # Solo contar años de matrícula >= año de cohorte
    merged = merged[merged["AÑO MATRICULA"] >= merged["AÑO COHORTE"]].copy()
    merged["ANIOS_DESDE_INGRESO"] = merged["AÑO MATRICULA"] - merged["AÑO COHORTE"]

    # Contar N presentes por carrera, cohorte y año
    presentes = (merged.groupby(["CODIGO SIGA", "AÑO COHORTE", "AÑO MATRICULA", "ANIOS_DESDE_INGRESO"])
                       ["RUT"].nunique().reset_index(name="N_PRESENTES"))

    # Unir con tamaño de cohorte y meta
    res = presentes.merge(cohortes, on=["CODIGO SIGA", "AÑO COHORTE"], how="left")
    res = res.merge(meta, on="CODIGO SIGA", how="left")
    res["CODIGO SIGA"] = res["CODIGO SIGA"].apply(
        lambda x: int(float(x)) if pd.notna(x) else None)
    return res.sort_values(
        ["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL", "AÑO COHORTE", "ANIOS_DESDE_INGRESO"]
    ).reset_index(drop=True)


def build_t5(df_via: pd.DataFrame, years: list) -> str:
    VIA_ORDER = ["Adm. Ordinaria DEMRE", "Adm. Especial (*)", "Complementaria", "Conducencia", "PACE"]
    IDX       = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"]

    # Normalizar por si quedó algún valor sin mapear
    REMAP = {"PACE": "PACE"}
    df_via = df_via.copy()
    df_via["VIA"] = df_via["VIA"].apply(lambda v: REMAP.get(str(v), str(v)))

    vias_pres = [v for v in VIA_ORDER if v in df_via["VIA"].values]
    n_sub     = len(vias_pres) + 1

    # Vía principal por carrera (una sola vez, fuera del loop)
    via_carr = (df_via.drop_duplicates(subset=["CODIGO SIGA","VIA"])
                .groupby("CODIGO SIGA")["VIA"]
                .first().to_dict())
    via_lbl_map = {
        "Adm. Ordinaria DEMRE": "DEMRE",
        "Adm. Especial (*)":    "Especial",
        "Complementaria":       "Compl.",
        "Conducencia":          "Conduc.",
        "PACE":                 "PACE",
    }

    # Construir lookup: (siga, year, via) -> N  y  (siga, year) -> Total
    lookup_n   = {}
    lookup_tot = {}
    for _, r in df_via.iterrows():
        k = (r["CODIGO SIGA"], int(r["AÑO INGRESO"]), r["VIA"])
        lookup_n[k] = int(r["N"])
        kt = (r["CODIGO SIGA"], int(r["AÑO INGRESO"]))
        if kt not in lookup_tot:
            lookup_tot[kt] = int(r["Total"])

    # Carreras únicas ordenadas
    carreras = (df_via[IDX]
                .drop_duplicates()
                .sort_values(["FACULTAD NOMPROPIO","NOMBRE CARRERA FINAL"])
                .reset_index(drop=True))

    # Totales UAH por año/vía
    tots = {}
    for y in years:
        yr = int(y)
        tots[yr] = {v: sum(lookup_n.get((s,yr,v),0)
                           for s in carreras["CODIGO SIGA"]) for v in vias_pres}
        tots[yr]["Total"] = sum(lookup_tot.get((s,yr),0)
                                for s in carreras["CODIGO SIGA"])

    h = ['<table class="t2 t2-via"><thead>']
    h.append('<tr class="rh1">')
    h.append('<th class="th-left" rowspan="2">FACULTAD</th>')
    h.append('<th class="th-left" rowspan="2">CÓD.</th>')
    h.append('<th class="th-left" rowspan="2">Vía ingreso</th>')
    h.append('<th class="th-left" rowspan="2" style="min-width:200px">CARRERA</th>')
    for y in years:
        h.append(f'<th class="th-yr" colspan="{n_sub}">AÑO DE INGRESO {int(y)}</th>')
    DISPLAY = {
        "Adm. Ordinaria DEMRE": "Admisión Ordinaria DEMRE",
        "Adm. Especial (*)":    "Admisión Especial (*)",
        "Complementaria":       "Complementaria",
        "Conducencia":          "Conducencia",
        "PACE":                 "PACE",
    }
    h.append('</tr><tr class="rh2">')
    for _ in years:
        for v in vias_pres:
            label = DISPLAY.get(v, v)
            if v == "PACE":
                h.append(f'<th class="sh-pace">{label}</th>')
            else:
                h.append(f'<th class="sh-pct">{label}</th>')
        h.append('<th class="sh-tot">Total</th>')
    h.append('</tr></thead><tbody>')

    prev_fac = None
    for _, row in carreras.iterrows():
        fac  = row["FACULTAD NOMPROPIO"]
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]
        via  = via_lbl_map.get(via_carr.get(siga,""),"")

        h.append('<tr>')
        if fac != prev_fac:
            h.append(f'<td class="c-fac">{fac}</td>')
        else:
            h.append('<td class="c-fac" style="color:transparent;border-top:none"></td>')
        prev_fac = fac
        h.append(f'<td class="c-siga">{int(siga) if pd.notna(siga) else ""}</td>')
        h.append(f'<td class="c-jorn">{via}</td>')
        h.append(f'<td class="c-carr" title="{carr}">{carr}</td>')

        for yi, y in enumerate(years):
            yr  = int(y)
            sep = " c-sep" if yi==0 else ""
            for vi, v in enumerate(vias_pres):
                val = lookup_n.get((siga,yr,v), 0)
                css = "sh-muj" if v=="Adm. Ordinaria DEMRE" else ("c-hom" if v=="PACE" else "c-pct")
                if vi==0: css += sep
                if val == 0:
                    h.append(f'<td class="c-empty {css}">—</td>')
                else:
                    h.append(f'<td class="{css} c-num">{val:,}</td>')
            tot = lookup_tot.get((siga,yr), 0)
            if tot == 0:
                h.append('<td class="c-empty c-tot">—</td>')
            else:
                h.append(f'<td class="c-tot c-num">{tot:,}</td>')
        h.append('</tr>')

    h.append('<tr class="tr-tot">')
    h.append('<td class="c-fac" colspan="3">TOTAL UAH</td>')
    h.append('<td class="c-carr"></td>')
    for yi, y in enumerate(years):
        yr = int(y)
        for vi, v in enumerate(vias_pres):
            sep = ' class="c-sep"' if vi==0 and yi==0 else ""
            h.append(f'<td{sep}>{tots[yr][v]:,}</td>')
        h.append(f'<td>{tots[yr]["Total"]:,}</td>')
    h.append('</tr></tbody></table>')
    return "".join(h)


def build_t3(df_ncs: pd.DataFrame, years: list) -> str:
    IDX = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"]

    df_n = df_ncs[df_ncs["TIPO"] == "NUEVO"].copy()
    df_s = df_ncs[df_ncs["TIPO"] == "SUPERIOR"].copy()

    def wide(ds, tipo):
        if ds.empty:
            return pd.DataFrame(columns=IDX)
        p = ds.pivot_table(index=IDX, columns="AÑO",
                           values=["N", "PCT", "Total"], aggfunc="first")
        p.columns = [f"{v}_{int(yr)}_{tipo}" for v, yr in p.columns]
        return p.reset_index()

    wn = wide(df_n, "N")
    ws = wide(df_s, "S")

    if wn.empty and ws.empty:
        return "<p>Sin datos</p>"
    elif wn.empty:
        mg = ws
    elif ws.empty:
        mg = wn
    else:
        mg = wn.merge(ws, on=IDX, how="outer")

    mg = mg.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)

    tots = {}
    for y in years:
        yr = int(y)
        nn = mg.get(f"N_{yr}_N", pd.Series([0])).fillna(0).sum()
        ns = mg.get(f"N_{yr}_S", pd.Series([0])).fillna(0).sum()
        t  = mg.get(f"Total_{yr}_N", pd.Series([0])).fillna(0).sum()
        tots[yr] = (int(nn), int(ns), int(t) if t else int(nn + ns))

    h = ['<table class="t2"><thead>']
    h.append('<tr class="rh1">')
    h.append('<th class="th-left" rowspan="2">FACULTAD</th>')
    h.append('<th class="th-left" rowspan="2">CÓD.</th>')
    h.append('<th class="th-left" rowspan="2" style="min-width:200px">CARRERA</th>')
    for y in years:
        h.append(f'<th class="th-yr" colspan="5">Año {int(y)}</th>')
    h.append('</tr><tr class="rh2">')
    for _ in years:
        h.append('<th class="sh-muj">Nuevo</th>')
        h.append('<th class="sh-hom">Curso Sup.</th>')
        h.append('<th class="sh-pct">% Nuevo</th>')
        h.append('<th class="sh-pct">% Curso Sup.</th>')
        h.append('<th class="sh-tot">Total</th>')
    h.append('</tr></thead><tbody>')

    prev_fac = None
    for _, row in mg.iterrows():
        fac  = row["FACULTAD NOMPROPIO"]
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]

        h.append('<tr>')
        if fac != prev_fac:
            h.append(f'<td class="c-fac">{fac}</td>')
        else:
            h.append('<td class="c-fac" style="color:transparent;border-top:none"></td>')
        prev_fac = fac
        h.append(f'<td class="c-siga">{int(siga) if pd.notna(siga) else ""}</td>')
        h.append(f'<td class="c-carr" title="{carr}">{carr}</td>')

        for yi, y in enumerate(years):
            yr  = int(y)
            sep = " c-sep" if yi == 0 else ""
            nn  = row.get(f"N_{yr}_N");   ns  = row.get(f"N_{yr}_S")
            pn  = row.get(f"PCT_{yr}_N"); ps  = row.get(f"PCT_{yr}_S")
            tot = row.get(f"Total_{yr}_N")

            if pd.isna(nn) and pd.isna(ns):
                h.append(f'<td class="c-empty{sep}" colspan="5">—</td>')
                continue

            nn_i  = int(nn)  if pd.notna(nn)  else 0
            ns_i  = int(ns)  if pd.notna(ns)  else 0
            tot_i = int(tot) if pd.notna(tot) else nn_i + ns_i
            pn_s  = f"{pn*100:.0f}%" if pd.notna(pn) else "—"
            ps_s  = f"{ps*100:.0f}%" if pd.notna(ps) else "—"

            h.append(f'<td class="c-muj{sep}">{nn_i:,}</td>')
            h.append(f'<td class="c-hom">{ns_i:,}</td>')
            h.append(f'<td class="c-pct">{pn_s}</td>')
            h.append(f'<td class="c-pct">{ps_s}</td>')
            h.append(f'<td class="c-tot">{tot_i:,}</td>')
        h.append('</tr>')

    h.append('<tr class="tr-tot">')
    h.append('<td class="c-fac" colspan="2">TOTAL UAH</td>')
    h.append('<td class="c-carr"></td>')
    for yi, y in enumerate(years):
        yr = int(y)
        nn_t, ns_t, tot_t = tots[yr]
        pn_t = f"{nn_t/tot_t*100:.0f}%" if tot_t else "—"
        ps_t = f"{ns_t/tot_t*100:.0f}%" if tot_t else "—"
        sep  = " c-sep" if yi == 0 else ""
        h.append(f'<td class="c-muj{sep}">{nn_t:,}</td>')
        h.append(f'<td class="c-hom">{ns_t:,}</td>')
        h.append(f'<td class="c-pct">{pn_t}</td>')
        h.append(f'<td class="c-pct">{ps_t}</td>')
        h.append(f'<td class="c-tot">{tot_t:,}</td>')
    h.append('</tr></tbody></table>')
    return "".join(h)


def _side():
    return Side(style="thin", color="D0D8E4")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

# Paleta UAH
AZ      = "1F3864"
AZ_MED  = "2F5496"
AZ_XCL  = "DEEAF1"
AZ_CL   = "BDD7EE"
BLANCO  = "FFFFFF"
GRIS_F  = "F8FAFC"
GRIS_C  = "FAFBFD"
ROSA    = "8B2252"
AZUL_H  = "1A4B8C"
VERDE   = "1A7A3C"
ROJO    = "C00000"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BLANCO, size=9, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def excel_t1(ws, df: pd.DataFrame, years: list, title: str):
    """Escribe tabla de carrera (tipo t1) en la hoja ws con formato UAH."""
    ycols = [y for y in years if y in df.columns]
    flat  = df[ycols].values.flatten()
    s_flat = pd.Series(flat).dropna()
    vmin, vmax = (s_flat.min(), s_flat.max()) if not s_flat.empty else (0, 1)

    # ── Fila título ──
    n_cols = 3 + len(ycols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(1, 1, title)
    tc.fill = _fill(AZ); tc.font = _font(bold=True, size=10)
    tc.alignment = _align("center")
    ws.row_dimensions[1].height = 18

    # ── Encabezados ──
    headers = ["FACULTAD", "CÓDIGO SIGA", "CARRERA"] + [int(y) for y in ycols]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.fill = _fill(AZ); c.font = _font(bold=True)
        c.alignment = _align("left" if ci <= 3 else "center")
        c.border = _border()
    ws.row_dimensions[2].height = 16

    # ── Datos ──
    prev_fac = None
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        fac  = row["FACULTAD NOMPROPIO"]
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]

        # Facultad
        c_fac = ws.cell(ri, 1, fac if fac != prev_fac else "")
        c_fac.fill = _fill(AZ_XCL)
        c_fac.font = Font(bold=True, color=AZ, size=8, name="Calibri")
        c_fac.alignment = _align("left")
        c_fac.border = _border()
        prev_fac = fac

        # Código
        c_sig = ws.cell(ri, 2, int(siga) if pd.notna(siga) else "")
        c_sig.fill = _fill(GRIS_F); c_sig.font = Font(bold=True, color=AZ_MED, size=8, name="Calibri")
        c_sig.alignment = _align("center"); c_sig.border = _border()

        # Carrera
        c_car = ws.cell(ri, 3, carr)
        c_car.fill = _fill(GRIS_C); c_car.font = Font(color="1E2D45", size=8, name="Calibri")
        c_car.alignment = _align("left"); c_car.border = _border()

        # Años
        for ci, y in enumerate(ycols, 4):
            v = row.get(y)
            c = ws.cell(ri, ci)
            if pd.isna(v) or v == 0:
                c.value = "—"; c.font = Font(color="CBD5E1", size=8, name="Calibri")
            else:
                c.value = int(v)
                # Heatmap
                t = max(0.0, min(1.0, (v - vmin) / (vmax - vmin))) if vmax != vmin else 0
                r_ = int(235 + t * (31  - 235))
                g_ = int(243 + t * (56  - 243))
                b_ = int(251 + t * (100 - 251))
                hex_bg = f"{r_:02X}{g_:02X}{b_:02X}"
                txt_col = BLANCO if t > 0.6 else "1E2D45"
                c.fill = _fill(hex_bg)
                c.font = Font(bold=True, color=txt_col, size=9, name="Calibri")
            c.alignment = _align("center"); c.border = _border()

        ws.row_dimensions[ri].height = 14

    # ── Total UAH ──
    tr = len(df) + 3
    ws.cell(tr, 1, "TOTAL UAH").fill = _fill(AZ_MED)
    ws.cell(tr, 1).font = _font(bold=True); ws.cell(tr, 1).border = _border()
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    ws.cell(tr, 3, "").fill = _fill(AZ_MED); ws.cell(tr, 3).border = _border()
    for ci, y in enumerate(ycols, 4):
        c = ws.cell(tr, ci, int(df[y].sum()) if y in df.columns else 0)
        c.fill = _fill(AZ_MED); c.font = _font(bold=True)
        c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[tr].height = 16

    # ── Anchos ──
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 38
    for ci in range(4, 4 + len(ycols)):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.freeze_panes = "D3"


def excel_t2(ws, df_sex: pd.DataFrame, years: list, title: str):
    """Escribe tabla de sexo (tipo t2) en la hoja ws con formato UAH."""
    IDX = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"]
    df_m = df_sex[df_sex["SEXO FINAL"] == "M"].copy()
    df_h = df_sex[df_sex["SEXO FINAL"] == "H"].copy()

    def wide(ds, sx):
        if ds.empty:
            return pd.DataFrame(columns=IDX)
        p = ds.pivot_table(index=IDX, columns="AÑO INGRESO",
                           values=["N", "PCT", "Total"], aggfunc="first")
        p.columns = [f"{v}_{int(yr)}_{sx}" for v, yr in p.columns]
        return p.reset_index()

    wm = wide(df_m, "M"); wh = wide(df_h, "H")
    if wm.empty and wh.empty: return
    mg = wm.merge(wh, on=IDX, how="outer") if not wm.empty and not wh.empty else (wm if not wm.empty else wh)
    mg = mg.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)

    n_cols = 4 + len(years) * 5
    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1, title).fill = _fill(AZ)
    ws.cell(1, 1).font = _font(bold=True, size=10)
    ws.cell(1, 1).alignment = _align("center")

    # Fila 1 headers
    for ci in range(1, 5):
        ws.cell(2, ci, ["FACULTAD","JORNADA","CÓDIGO SIGA","CARRERA"][ci-1])
        ws.cell(2, ci).fill = _fill(AZ); ws.cell(2, ci).font = _font(bold=True)
        ws.cell(2, ci).alignment = _align("left" if ci in (1,4) else "center")
        ws.cell(2, ci).border = _border()
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)

    for yi, y in enumerate(years):
        base = 5 + yi * 5
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base+4)
        hc = ws.cell(2, base, f"Año {int(y)}")
        hc.fill = _fill(AZ_MED); hc.font = _font(bold=True)
        hc.alignment = _align("center"); hc.border = _border()
        for ci2, lbl in enumerate(["Mujeres","Hombres","% Muj","% Hom","Total"]):
            c = ws.cell(3, base+ci2, lbl)
            c.fill = _fill("2A4070"); c.font = _font(bold=True, size=8)
            c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 14

    jornada_map = df_sex.groupby("CODIGO SIGA")["JORNADA"].first().to_dict()
    prev_fac = None
    for ri, (_, row) in enumerate(mg.iterrows(), 4):
        fac  = row["FACULTAD NOMPROPIO"]
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]
        jorn = jornada_map.get(siga, "")

        c1 = ws.cell(ri, 1, fac if fac != prev_fac else "")
        c1.fill = _fill(AZ_XCL); c1.font = Font(bold=True, color=AZ, size=8, name="Calibri")
        c1.alignment = _align("left"); c1.border = _border()
        prev_fac = fac

        ws.cell(ri, 2, jorn).fill = _fill("F0F4F8")
        ws.cell(ri, 2).font = Font(color="4A6080", size=8, name="Calibri")
        ws.cell(ri, 2).alignment = _align("center"); ws.cell(ri, 2).border = _border()

        ws.cell(ri, 3, int(siga) if pd.notna(siga) else "").fill = _fill(GRIS_F)
        ws.cell(ri, 3).font = Font(bold=True, color=AZ_MED, size=8, name="Calibri")
        ws.cell(ri, 3).alignment = _align("center"); ws.cell(ri, 3).border = _border()

        ws.cell(ri, 4, carr).fill = _fill(GRIS_C)
        ws.cell(ri, 4).font = Font(color="1E2D45", size=8, name="Calibri")
        ws.cell(ri, 4).alignment = _align("left"); ws.cell(ri, 4).border = _border()

        for yi, y in enumerate(years):
            yr   = int(y)
            base = 5 + yi * 5
            nm   = row.get(f"N_{yr}_M");   nh  = row.get(f"N_{yr}_H")
            pm   = row.get(f"PCT_{yr}_M"); ph  = row.get(f"PCT_{yr}_H")
            tot  = row.get(f"Total_{yr}_M")
            nm_i = int(nm) if pd.notna(nm) else 0
            nh_i = int(nh) if pd.notna(nh) else 0
            tot_i = int(tot) if pd.notna(tot) else nm_i + nh_i
            pm_s = f"{pm*100:.0f}%" if pd.notna(pm) else "—"
            ph_s = f"{ph*100:.0f}%" if pd.notna(ph) else "—"
            vals = [nm_i, nh_i, pm_s, ph_s, tot_i]
            cols_c = [ROSA, AZUL_H, "556677", "556677", AZ]
            for ci2, (val, col) in enumerate(zip(vals, cols_c)):
                c = ws.cell(ri, base+ci2, val)
                c.font = Font(bold=(ci2 in (0,1,4)), color=col, size=8, name="Calibri")
                c.alignment = _align("center"); c.border = _border()
                if ci2 == 4:
                    c.fill = _fill("F0F5FA")
        ws.row_dimensions[ri].height = 14

    # Total
    tr = len(mg) + 4
    for ci in range(1, 4):
        c = ws.cell(tr, ci, "TOTAL UAH" if ci == 1 else "")
        c.fill = _fill(AZ_MED); c.font = _font(bold=True)
        c.border = _border()
    ws.cell(tr, 4, "").fill = _fill(AZ); ws.cell(tr, 4).border = _border()
    for yi, y in enumerate(years):
        yr   = int(y)
        base = 5 + yi * 5
        nm_t = int(mg.get(f"N_{yr}_M", pd.Series([0])).fillna(0).sum())
        nh_t = int(mg.get(f"N_{yr}_H", pd.Series([0])).fillna(0).sum())
        tot_t = nm_t + nh_t
        pm_t = f"{nm_t/tot_t*100:.0f}%" if tot_t else "—"
        ph_t = f"{nh_t/tot_t*100:.0f}%" if tot_t else "—"
        for ci2, val in enumerate([nm_t, nh_t, pm_t, ph_t, tot_t]):
            c = ws.cell(tr, base+ci2, val)
            c.fill = _fill(AZ_MED); c.font = _font(bold=True)
            c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[tr].height = 16

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 34
    for ci in range(5, 5 + len(years)*5):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.freeze_panes = "E4"


def excel_t3(ws, df_ncs: pd.DataFrame, years: list, title: str):
    """Escribe tabla nueva matrícula/cursos superiores en la hoja ws."""
    IDX = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"]
    df_n = df_ncs[df_ncs["TIPO"] == "NUEVO"].copy()
    df_s = df_ncs[df_ncs["TIPO"] == "SUPERIOR"].copy()

    def wide(ds, tipo):
        if ds.empty: return pd.DataFrame(columns=IDX)
        p = ds.pivot_table(index=IDX, columns="AÑO", values=["N","PCT","Total"], aggfunc="first")
        p.columns = [f"{v}_{int(yr)}_{tipo}" for v, yr in p.columns]
        return p.reset_index()

    wn = wide(df_n, "N"); ws2 = wide(df_s, "S")
    if wn.empty and ws2.empty: return
    mg = wn.merge(ws2, on=IDX, how="outer") if not wn.empty and not ws2.empty else (wn if not wn.empty else ws2)
    mg = mg.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)

    n_cols = 3 + len(years) * 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1, title).fill = _fill(AZ)
    ws.cell(1, 1).font = _font(bold=True, size=10)
    ws.cell(1, 1).alignment = _align("center")

    for ci, lbl in enumerate(["FACULTAD","CÓDIGO SIGA","CARRERA"], 1):
        c = ws.cell(2, ci, lbl)
        c.fill = _fill(AZ); c.font = _font(bold=True)
        c.alignment = _align("left" if ci != 2 else "center")
        c.border = _border()
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)

    for yi, y in enumerate(years):
        base = 4 + yi * 5
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base+4)
        hc = ws.cell(2, base, f"Año {int(y)}")
        hc.fill = _fill(AZ_MED); hc.font = _font(bold=True)
        hc.alignment = _align("center"); hc.border = _border()
        for ci2, lbl in enumerate(["Nuevo","Curso Sup.","% Nuevo","% Curso Sup.","Total"]):
            c = ws.cell(3, base+ci2, lbl)
            c.fill = _fill("2A4070"); c.font = _font(bold=True, size=8)
            c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 14

    prev_fac = None
    for ri, (_, row) in enumerate(mg.iterrows(), 4):
        fac  = row["FACULTAD NOMPROPIO"]
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]

        c1 = ws.cell(ri, 1, fac if fac != prev_fac else "")
        c1.fill = _fill(AZ_XCL); c1.font = Font(bold=True, color=AZ, size=8, name="Calibri")
        c1.alignment = _align("left"); c1.border = _border(); prev_fac = fac

        ws.cell(ri, 2, int(siga) if pd.notna(siga) else "").fill = _fill(GRIS_F)
        ws.cell(ri, 2).font = Font(bold=True, color=AZ_MED, size=8, name="Calibri")
        ws.cell(ri, 2).alignment = _align("center"); ws.cell(ri, 2).border = _border()

        ws.cell(ri, 3, carr).fill = _fill(GRIS_C)
        ws.cell(ri, 3).font = Font(color="1E2D45", size=8, name="Calibri")
        ws.cell(ri, 3).alignment = _align("left"); ws.cell(ri, 3).border = _border()

        for yi, y in enumerate(years):
            yr   = int(y)
            base = 4 + yi * 5
            nn   = row.get(f"N_{yr}_N");   ns  = row.get(f"N_{yr}_S")
            pn   = row.get(f"PCT_{yr}_N"); ps  = row.get(f"PCT_{yr}_S")
            tot  = row.get(f"Total_{yr}_N")
            nn_i = int(nn) if pd.notna(nn) else 0
            ns_i = int(ns) if pd.notna(ns) else 0
            tot_i = int(tot) if pd.notna(tot) else nn_i + ns_i
            pn_s = f"{pn*100:.0f}%" if pd.notna(pn) else "—"
            ps_s = f"{ps*100:.0f}%" if pd.notna(ps) else "—"
            cols_c = [ROSA, AZUL_H, "556677", "556677", AZ]
            for ci2, (val, col) in enumerate(zip([nn_i,ns_i,pn_s,ps_s,tot_i], cols_c)):
                c = ws.cell(ri, base+ci2, val)
                c.font = Font(bold=(ci2 in (0,1,4)), color=col, size=8, name="Calibri")
                c.alignment = _align("center"); c.border = _border()
                if ci2 == 4: c.fill = _fill("F0F5FA")
        ws.row_dimensions[ri].height = 14

    tr = len(mg) + 4
    for ci in range(1, 3):
        c = ws.cell(tr, ci, "TOTAL UAH" if ci == 1 else "")
        c.fill = _fill(AZ_MED); c.font = _font(bold=True); c.border = _border()
    ws.cell(tr, 3, "").fill = _fill(AZ); ws.cell(tr, 3).border = _border()
    for yi, y in enumerate(years):
        yr   = int(y); base = 4 + yi * 5
        nn_t = int(mg.get(f"N_{yr}_N", pd.Series([0])).fillna(0).sum())
        ns_t = int(mg.get(f"N_{yr}_S", pd.Series([0])).fillna(0).sum())
        tot_t = nn_t + ns_t
        pn_t = f"{nn_t/tot_t*100:.0f}%" if tot_t else "—"
        ps_t = f"{ns_t/tot_t*100:.0f}%" if tot_t else "—"
        for ci2, val in enumerate([nn_t,ns_t,pn_t,ps_t,tot_t]):
            c = ws.cell(tr, base+ci2, val)
            c.fill = _fill(AZ_MED); c.font = _font(bold=True)
            c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[tr].height = 16

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 34
    for ci in range(4, 4 + len(years)*5):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    ws.freeze_panes = "D4"


def excel_t5(ws, df_via: pd.DataFrame, years: list, title: str):
    """Escribe tabla de vía de ingreso en la hoja ws con formato UAH."""
    VIA_ORDER  = ["Adm. Ordinaria DEMRE", "Adm. Especial (*)", "Complementaria", "Conducencia", "PACE"]
    VIA_LABELS = {
        "Adm. Ordinaria DEMRE": "Adm. Ordinaria DEMRE",
        "Adm. Especial (*)":    "Adm. Especial (*)",
        "Complementaria":       "Complementaria",
        "Conducencia":          "Conducencia",
        "PACE":                 "PACE",
    }
    IDX = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"]

    vias_pres = [v for v in VIA_ORDER if v in df_via["VIA"].values]
    n_sub     = len(vias_pres) + 1  # vías + Total

    # Meta canónica por código
    via_carr = (df_via.drop_duplicates(subset=["CODIGO SIGA", "VIA"])
                .groupby("CODIGO SIGA")["VIA"].first().to_dict())

    # Lookup (siga, year, via) -> N y (siga, year) -> Total
    lookup_n   = {}
    lookup_tot = {}
    for _, r in df_via.iterrows():
        lookup_n[(r["CODIGO SIGA"], int(r["AÑO INGRESO"]), r["VIA"])] = int(r["N"])
        kt = (r["CODIGO SIGA"], int(r["AÑO INGRESO"]))
        if kt not in lookup_tot:
            lookup_tot[kt] = int(r["Total"])

    carreras = (df_via[IDX].drop_duplicates()
                .sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"])
                .reset_index(drop=True))

    n_cols = 4 + len(years) * n_sub
    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.cell(1, 1, title).fill = _fill(AZ)
    ws.cell(1, 1).font = _font(bold=True, size=10)
    ws.cell(1, 1).alignment = _align("center")
    ws.row_dimensions[1].height = 18

    # Fila 1 headers — cols fijas
    for ci, lbl in enumerate(["FACULTAD", "CÓD.", "Vía ingreso", "CARRERA"], 1):
        c = ws.cell(2, ci, lbl)
        c.fill = _fill(AZ); c.font = _font(bold=True)
        c.alignment = _align("left" if ci in (1, 4) else "center")
        c.border = _border()
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)

    # Fila 1 — años
    for yi, y in enumerate(years):
        base = 5 + yi * n_sub
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base + n_sub - 1)
        hc = ws.cell(2, base, f"AÑO DE INGRESO {int(y)}")
        hc.fill = _fill(AZ_MED); hc.font = _font(bold=True)
        hc.alignment = _align("center"); hc.border = _border()
        # Fila 2 — sub-headers vías
        for vi, v in enumerate(vias_pres):
            c = ws.cell(3, base + vi, VIA_LABELS.get(v, v))
            c.fill = _fill("2A4070"); c.font = _font(bold=False, color="C8D8E8", size=8)
            c.alignment = _align("center"); c.border = _border()
        c = ws.cell(3, base + len(vias_pres), "Total")
        c.fill = _fill("1F3864"); c.font = _font(bold=True, size=8)
        c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 14

    # Totales UAH
    tots = {}
    for y in years:
        yr = int(y)
        tots[yr] = {v: sum(lookup_n.get((s, yr, v), 0) for s in carreras["CODIGO SIGA"])
                    for v in vias_pres}
        tots[yr]["Total"] = sum(lookup_tot.get((s, yr), 0) for s in carreras["CODIGO SIGA"])

    via_lbl_map = {
        "Adm. Ordinaria DEMRE": "DEMRE",
        "Adm. Especial (*)":    "Especial",
        "Complementaria":       "Compl.",
        "Conducencia":          "Conduc.",
        "PACE":                 "PACE",
    }

    prev_fac = None
    for ri, (_, row) in enumerate(carreras.iterrows(), 4):
        fac  = row["FACULTAD NOMPROPIO"]
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]
        via  = via_lbl_map.get(via_carr.get(siga, ""), "")

        c1 = ws.cell(ri, 1, fac if fac != prev_fac else "")
        c1.fill = _fill(AZ_XCL); c1.font = Font(bold=True, color=AZ, size=8, name="Calibri")
        c1.alignment = _align("left"); c1.border = _border(); prev_fac = fac

        ws.cell(ri, 2, int(siga) if pd.notna(siga) else "").fill = _fill(GRIS_F)
        ws.cell(ri, 2).font = Font(bold=True, color=AZ_MED, size=8, name="Calibri")
        ws.cell(ri, 2).alignment = _align("center"); ws.cell(ri, 2).border = _border()

        ws.cell(ri, 3, via).fill = _fill("F0F4F8")
        ws.cell(ri, 3).font = Font(color="4A6080", size=8, name="Calibri")
        ws.cell(ri, 3).alignment = _align("center"); ws.cell(ri, 3).border = _border()

        ws.cell(ri, 4, carr).fill = _fill(GRIS_C)
        ws.cell(ri, 4).font = Font(color="1E2D45", size=8, name="Calibri")
        ws.cell(ri, 4).alignment = _align("left"); ws.cell(ri, 4).border = _border()

        for yi, y in enumerate(years):
            yr   = int(y)
            base = 5 + yi * n_sub
            for vi, v in enumerate(vias_pres):
                val = lookup_n.get((siga, yr, v), 0)
                c = ws.cell(ri, base + vi, val if val > 0 else "—")
                c.font = Font(color=(AZ if val > 0 else "CBD5E1"), bold=(val > 0), size=8, name="Calibri")
                c.alignment = _align("center"); c.border = _border()
            tot = lookup_tot.get((siga, yr), 0)
            c = ws.cell(ri, base + len(vias_pres), tot if tot > 0 else "—")
            c.fill = _fill("F0F5FA")
            c.font = Font(bold=True, color=(AZ if tot > 0 else "CBD5E1"), size=8, name="Calibri")
            c.alignment = _align("center"); c.border = _border()
        ws.row_dimensions[ri].height = 14

    # Fila total UAH
    tr = len(carreras) + 4
    for ci, lbl in enumerate(["TOTAL UAH", "", "", ""], 1):
        c = ws.cell(tr, ci, lbl)
        c.fill = _fill(AZ_MED if ci > 1 else AZ)
        c.font = _font(bold=True); c.border = _border()
    for yi, y in enumerate(years):
        yr   = int(y)
        base = 5 + yi * n_sub
        for vi, v in enumerate(vias_pres):
            c = ws.cell(tr, base + vi, tots[yr][v])
            c.fill = _fill(AZ_MED); c.font = _font(bold=True)
            c.alignment = _align("center"); c.border = _border()
        c = ws.cell(tr, base + len(vias_pres), tots[yr]["Total"])
        c.fill = _fill(AZ); c.font = _font(bold=True)
        c.alignment = _align("center"); c.border = _border()
    ws.row_dimensions[tr].height = 16

    # Anchos
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 34
    for ci in range(5, 5 + len(years) * n_sub):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.freeze_panes = "E4"


def build_t4(df_ret: pd.DataFrame, cohortes: list) -> str:
    """Tabla de retención 1er año por carrera y cohorte."""
    IDX = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"]

    pivot = df_ret.pivot_table(
        index=IDX, columns="AÑO COHORTE", values="PCT", aggfunc="first"
    ).reset_index()
    pivot.columns.name = None
    pivot = pivot.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)

    cols_coh = [c for c in cohortes if c in pivot.columns]
    # Últimas 3 cohortes con datos para promedio (excluye la más reciente que aún no tiene año+1)
    last3 = cols_coh[-3:] if len(cols_coh) >= 3 else cols_coh

    # Diferencia entre las 2 últimas cohortes disponibles
    show_dif = len(cols_coh) >= 2
    c_ant, c_ult = (cols_coh[-2], cols_coh[-1]) if show_dif else (None, None)

    h = ['<table class="t1"><thead><tr>']
    h += ['<th class="c-fac">FACULTAD</th>',
          '<th class="c-siga">CÓDIGO</th>',
          '<th class="c-carrera">CARRERA</th>']
    for c in cols_coh:
        h.append(f'<th>Coh. {int(c)}</th>')
    if show_dif:
        h.append(f'<th class="c-diff">Dif {int(c_ant)}–{int(c_ult)}</th>')
    h.append(f'<th class="c-var">Prom. últ. 3 Cohortes</th>')
    h.append('</tr></thead><tbody>')

    prev_fac = None
    for _, row in pivot.iterrows():
        fac  = row["FACULTAD NOMPROPIO"]
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]
        h.append('<tr>')
        if fac != prev_fac:
            h.append(f'<td class="c-fac">{fac}</td>')
        else:
            h.append('<td class="c-fac" style="color:transparent;border-top:none"></td>')
        prev_fac = fac
        h.append(f'<td class="c-siga">{int(siga) if pd.notna(siga) else ""}</td>')
        h.append(f'<td class="c-carrera">{carr}</td>')

        for c in cols_coh:
            v = row.get(c)
            if pd.isna(v):
                h.append('<td class="c-empty c-num">—</td>')
            else:
                pct = v * 100
                # Color por nivel: verde > 80%, amarillo 60–80%, rojo < 60%
                if pct >= 80:
                    bg = "background:#E8F5E9;color:#1A7A3C;"
                elif pct >= 60:
                    bg = "background:#FFF8E1;color:#7A5C00;"
                else:
                    bg = "background:#FFEBEE;color:#C00000;"
                h.append(f'<td class="c-num" style="{bg}font-weight:700">{pct:.0f}%</td>')

        if show_dif:
            va = row.get(c_ant); vu = row.get(c_ult)
            if pd.isna(va) or pd.isna(vu):
                h.append('<td class="c-empty c-num">—</td>')
            else:
                d = (vu - va) * 100
                sign = "+" if d > 0 else ""
                color = "color:#1A7A3C;font-weight:700" if d > 0 else ("color:#C00000;font-weight:700" if d < 0 else "")
                h.append(f'<td class="c-num" style="{color}">{sign}{d:.0f}%</td>')

        vals_prom = [row.get(c) for c in last3 if pd.notna(row.get(c))]
        if vals_prom:
            prom = sum(vals_prom) / len(vals_prom) * 100
            if prom >= 80:
                bg = "background:#C8E6C9;color:#1A7A3C;"
            elif prom >= 60:
                bg = "background:#FFF9C4;color:#7A5C00;"
            else:
                bg = "background:#FFCDD2;color:#C00000;"
            h.append(f'<td class="c-num" style="{bg}font-weight:700">{prom:.0f}%</td>')
        else:
            h.append('<td class="c-empty c-num">—</td>')
        h.append('</tr>')

    # Total UAH (promedio ponderado por cohorte)
    h.append('<tr class="tr-tot"><td class="c-fac" colspan="2">TOTAL UAH</td><td class="c-carrera"></td>')
    tot_df = df_ret.groupby("AÑO COHORTE").agg(
        N_RETENIDOS=("N_RETENIDOS", "sum"), N_COHORTE=("N_COHORTE", "sum")
    ).reset_index()
    tot_dict = {r["AÑO COHORTE"]: r["N_RETENIDOS"]/r["N_COHORTE"] for _, r in tot_df.iterrows() if r["N_COHORTE"] > 0}
    for c in cols_coh:
        v = tot_dict.get(c)
        h.append(f'<td>{v*100:.0f}%</td>' if v is not None else '<td>—</td>')
    if show_dif:
        va = tot_dict.get(c_ant); vu = tot_dict.get(c_ult)
        if va and vu:
            d = (vu - va) * 100
            sign = "+" if d > 0 else ""
            h.append(f'<td>{sign}{d:.0f}%</td>')
        else:
            h.append('<td>—</td>')
    prom_last3 = [tot_dict[c] for c in last3 if c in tot_dict]
    prom_uah = sum(prom_last3)/len(prom_last3)*100 if prom_last3 else None
    h.append(f'<td>{prom_uah:.0f}%</td>' if prom_uah else '<td>—</td>')
    h.append('</tr></tbody></table>')
    return "".join(h)


def heatmap_color(val, vmin, vmax):
    if pd.isna(val) or vmax == vmin:
        return ""
    t = max(0.0, min(1.0, (val - vmin) / (vmax - vmin)))
    r = int(235 + t * (31  - 235))
    g = int(243 + t * (56  - 243))
    b = int(251 + t * (100 - 251))
    text = "white" if t > 0.6 else "#1E2D45"
    return f"background:rgb({r},{g},{b});color:{text};"


def dl_link(data: bytes, filename: str, label: str, mime: str) -> str:
    b64 = base64.b64encode(data).decode()
    return f'<a class="dl-btn" href="data:{mime};base64,{b64}" download="{filename}">{label}</a>'


def build_t1(df: pd.DataFrame, years: list, heatmap: bool, show_diff: bool = False) -> str:
    ycols = [y for y in years if y in df.columns]
    flat  = df[ycols].values.flatten()
    vmin, vmax = pd.Series(flat).dropna().min(), pd.Series(flat).dropna().max()

    # Calcular diferencia y variación entre los 2 últimos años disponibles
    diff_col = var_col = None
    if show_diff and len(ycols) >= 2:
        y_ant, y_ult = ycols[-2], ycols[-1]
        diff_col = f"_diff"
        var_col  = f"_var"
        df = df.copy()
        v_ant = df[y_ant].fillna(0)
        v_ult = df[y_ult].fillna(0)
        df[diff_col] = v_ult - v_ant
        df[var_col]  = df.apply(
            lambda r: (r[y_ult] - r[y_ant]) / r[y_ant] if pd.notna(r[y_ant]) and r[y_ant] != 0 else None,
            axis=1
        )

    h = ['<table class="t1"><thead><tr>']
    h += ['<th class="c-fac">FACULTAD</th>',
          '<th class="c-siga">CÓDIGO SIGA</th>',
          '<th class="c-carrera">CARRERA</th>']
    for y in ycols:
        h.append(f'<th>{int(y)}</th>')
    if show_diff and diff_col:
        h.append(f'<th class="c-diff">Dif. {int(y_ant)}–{int(y_ult)}</th>')
        h.append(f'<th class="c-var">Var. %</th>')
    h.append('</tr></thead><tbody>')

    prev = None
    for _, row in df.iterrows():
        fac, siga, carr = row["FACULTAD NOMPROPIO"], row["CODIGO SIGA"], row["NOMBRE CARRERA FINAL"]
        h.append('<tr>')
        if fac != prev:
            h.append(f'<td class="c-fac">{fac}</td>')
        else:
            h.append('<td class="c-fac" style="color:transparent;border-top:none"></td>')
        prev = fac
        h.append(f'<td class="c-siga">{int(siga) if pd.notna(siga) else ""}</td>')
        h.append(f'<td class="c-carrera">{carr}</td>')
        for y in ycols:
            v = row.get(y)
            if pd.isna(v) or v == 0:
                h.append('<td class="c-empty c-num">—</td>')
            else:
                s = heatmap_color(v, vmin, vmax) if heatmap else ""
                h.append(f'<td class="c-num" style="{s}">{int(v)}</td>')
        if show_diff and diff_col:
            d = row.get(diff_col)
            v = row.get(var_col)
            if pd.isna(d):
                h.append('<td class="c-empty c-num">—</td>')
                h.append('<td class="c-empty c-num">—</td>')
            else:
                d_i   = int(d)
                sign  = "+" if d_i > 0 else ""
                color = "color:#1A7A3C;font-weight:700" if d_i > 0 else ("color:#C00000;font-weight:700" if d_i < 0 else "")
                h.append(f'<td class="c-num" style="{color}">{sign}{d_i:,}</td>')
                if pd.notna(v):
                    pct   = v * 100
                    psign = "+" if pct > 0 else ""
                    h.append(f'<td class="c-num" style="{color}">{psign}{pct:.1f}%</td>')
                else:
                    h.append('<td class="c-empty c-num">—</td>')
        h.append('</tr>')

    h.append('<tr class="tr-tot"><td class="c-fac" colspan="2">TOTAL UAH</td>'
             '<td class="c-carrera"></td>')
    for y in ycols:
        h.append(f'<td>{int(df[y].sum()) if y in df.columns else 0:,}</td>')
    if show_diff and diff_col:
        tot_ant  = df[y_ant].fillna(0).sum()
        tot_ult  = df[y_ult].fillna(0).sum()
        tot_d    = int(tot_ult - tot_ant)
        tot_sign = "+" if tot_d > 0 else ""
        tot_col  = "color:white;font-weight:700"
        tot_pct  = f"{tot_sign}{(tot_d/tot_ant*100):.1f}%" if tot_ant else "—"
        h.append(f'<td style="{tot_col}">{tot_sign}{tot_d:,}</td>')
        h.append(f'<td style="{tot_col}">{tot_pct}</td>')
    h.append('</tr></tbody></table>')
    return "".join(h)


def build_t2(df_sex: pd.DataFrame, years: list) -> str:
    IDX = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"]

    df_m = df_sex[df_sex["SEXO FINAL"] == "M"].copy()
    df_h = df_sex[df_sex["SEXO FINAL"] == "H"].copy()

    def wide(ds, sx):
        p = ds.pivot_table(index=IDX, columns="AÑO INGRESO",
                           values=["N", "PCT", "Total"], aggfunc="first")
        p.columns = [f"{v}_{int(yr)}_{sx}" for v, yr in p.columns]
        return p.reset_index()

    jornada_carr = df_sex.groupby("CODIGO SIGA")["JORNADA"].first().reset_index()

    mg = wide(df_m, "M").merge(wide(df_h, "H"), on=IDX, how="outer")
    mg = mg.merge(jornada_carr, on="CODIGO SIGA", how="left")
    mg = mg.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)

    tots = {}
    for y in years:
        nm = mg.get(f"N_{int(y)}_M", pd.Series([0])).sum()
        nh = mg.get(f"N_{int(y)}_H", pd.Series([0])).sum()
        tots[y] = (int(nm), int(nh), int(nm + nh))

    h = ['<table class="t2"><thead>']
    h.append('<tr class="rh1">')
    h.append('<th class="th-left" rowspan="2">FACULTAD</th>')
    h.append('<th class="th-left" rowspan="2">JORNADA</th>')
    h.append('<th class="th-left" rowspan="2">CÓDIGO SIGA</th>')
    h.append('<th class="th-left" rowspan="2" style="min-width:200px">CARRERA</th>')
    for y in years:
        h.append(f'<th class="th-yr" colspan="5">Año {int(y)}</th>')
    h.append('</tr>')

    h.append('<tr class="rh2">')
    for _ in years:
        h.append('<th class="sh-muj">Mujeres</th>')
        h.append('<th class="sh-hom">Hombres</th>')
        h.append('<th class="sh-pct">% Mujeres</th>')
        h.append('<th class="sh-pct">% Hombres</th>')
        h.append('<th class="sh-tot">Total</th>')
    h.append('</tr></thead><tbody>')

    prev_fac = None
    for _, row in mg.iterrows():
        fac  = row["FACULTAD NOMPROPIO"]
        jorn = row.get("JORNADA", "")
        siga = row["CODIGO SIGA"]
        carr = row["NOMBRE CARRERA FINAL"]

        h.append('<tr>')
        if fac != prev_fac:
            h.append(f'<td class="c-fac">{fac}</td>')
        else:
            h.append('<td class="c-fac" style="color:transparent;border-top:none"></td>')
        prev_fac = fac

        h.append(f'<td class="c-jorn">{jorn}</td>')
        h.append(f'<td class="c-siga">{int(siga) if pd.notna(siga) else ""}</td>')
        h.append(f'<td class="c-carr" title="{carr}">{carr}</td>')

        for yi, y in enumerate(years):
            yr   = int(y)
            sep  = " c-sep" if yi == 0 else ""
            nm   = row.get(f"N_{yr}_M");   nh  = row.get(f"N_{yr}_H")
            pm   = row.get(f"PCT_{yr}_M"); ph  = row.get(f"PCT_{yr}_H")
            tot  = row.get(f"Total_{yr}_M")

            if pd.isna(nm) and pd.isna(nh):
                h.append(f'<td class="c-empty{sep}" colspan="5">—</td>')
                continue

            nm_i  = int(nm)  if pd.notna(nm)  else 0
            nh_i  = int(nh)  if pd.notna(nh)  else 0
            tot_i = int(tot) if pd.notna(tot) else nm_i + nh_i
            pm_s  = f"{pm*100:.0f}%" if pd.notna(pm) else "—"
            ph_s  = f"{ph*100:.0f}%" if pd.notna(ph) else "—"

            h.append(f'<td class="c-muj{sep}">{nm_i}</td>')
            h.append(f'<td class="c-hom">{nh_i}</td>')
            h.append(f'<td class="c-pct">{pm_s}</td>')
            h.append(f'<td class="c-pct">{ph_s}</td>')
            h.append(f'<td class="c-tot">{tot_i}</td>')
        h.append('</tr>')

    h.append('<tr class="tr-tot">')
    h.append('<td class="c-fac" colspan="3">TOTAL UAH</td>')
    h.append('<td class="c-carr"></td>')
    for yi, y in enumerate(years):
        nm_t, nh_t, tot_t = tots[y]
        pm_t = f"{nm_t/tot_t*100:.0f}%" if tot_t else "—"
        ph_t = f"{nh_t/tot_t*100:.0f}%" if tot_t else "—"
        sep  = " c-sep" if yi == 0 else ""
        h.append(f'<td class="c-muj{sep}">{nm_t:,}</td>')
        h.append(f'<td class="c-hom">{nh_t:,}</td>')
        h.append(f'<td class="c-pct">{pm_t}</td>')
        h.append(f'<td class="c-pct">{ph_t}</td>')
        h.append(f'<td class="c-tot">{tot_t:,}</td>')
    h.append('</tr></tbody></table>')
    return "".join(h)


# ──────────────────────────────────────────────────────────────────────────────
# HEADER
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="uah-header">
  <div class="uah-logo">UAH</div>
  <div>
    <div class="uah-title">UAH en Cifras</div>
    <div class="uah-desc">Admisión Pregrado · Matrícula Completa</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# PESTAÑAS PRINCIPALES
# ──────────────────────────────────────────────────────────────────────────────
main_adm, main_mat, main_oferta = st.tabs([
    "🎓  Admisión",
    "📋  Matrícula Completa",
    "📚  Oferta Académica",
])

# ══════════════════════════════════════════════════════════════════════════════
# PESTAÑA PRINCIPAL: ADMISIÓN
# ══════════════════════════════════════════════════════════════════════════════
with main_adm:

    file_bytes, hoja_sel = selector_origen_archivo(
        nombre="Admisión",
        ejemplo="1. ADM PREG 2011 2026 30 DE ABRIL.xlsx",
        key_prefix="adm",
        archivo_enc="admision.enc",
    )

    if not file_bytes:
        st.info("⬆️  Sube un archivo Excel para comenzar.")
    else:
        st.markdown(f'<div class="sheet-badge">🗂 &nbsp;Hoja activa: <strong>{hoja_sel}</strong></div>',
                    unsafe_allow_html=True)

        tab1, tab2, tab3, tab4 = st.tabs([
            "📊  Admisión por Carrera",
            "👥  Admisión por Sexo",
            "🛤️  Admisión por Vía de Ingreso",
            "✅  Admisión Regular",
        ])

        with tab1:
            try:
                with st.spinner("Procesando…"):
                    df_adm = procesar_admision(file_bytes, hoja_sel)
            except Exception as e:
                st.error(f"No se pudo leer la hoja.\n\n`{e}`"); st.stop()

            year_all = sorted([c for c in df_adm.columns
                               if isinstance(c, (int, float)) and 2000 <= float(c) <= 2030])
            if not year_all:
                st.warning("La hoja no contiene columnas de año. Prueba con BASE 1ER SEM."); st.stop()

            yr_last      = year_all[-1]
            total_ultimo = int(df_adm[yr_last].sum()) if yr_last in df_adm.columns else 0
            total_hist   = int(sum(df_adm[y].sum() for y in year_all if y in df_adm.columns))

            st.markdown(f"""
            <div class="metric-strip">
              <div class="metric-card"><div class="mc-label">Carreras</div>
                <div class="mc-value">{int((df_adm[yr_last] > 0).sum())}</div></div>
              <div class="metric-card"><div class="mc-label">Facultades</div>
                <div class="mc-value">{df_adm["FACULTAD NOMPROPIO"].nunique()}</div></div>
              <div class="metric-card"><div class="mc-label">Admitidos {int(yr_last)}</div>
                <div class="mc-value">{total_ultimo:,}</div></div>
              <div class="metric-card accent"><div class="mc-label">Total histórico</div>
                <div class="mc-value">{total_hist:,}</div></div>
              <div class="metric-card"><div class="mc-label">Años</div>
                <div class="mc-value">{int(year_all[0])}–{int(year_all[-1])}</div></div>
            </div>""", unsafe_allow_html=True)

            f1, f2, f3, f4 = st.columns([2, 2, 2, 1])
            with f1:
                facs    = sorted(df_adm["FACULTAD NOMPROPIO"].dropna().unique())
                fac_sel = st.multiselect("Facultad", facs, default=facs, placeholder="Todas")
            with f2:
                rango = st.select_slider("Rango de años",
                                         options=[int(y) for y in year_all],
                                         value=(int(year_all[0]), int(year_all[-1])))
            with f3:
                busqueda = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…")
            with f4:
                heatmap = st.toggle("Mapa de calor", value=True)

            df_f = df_adm.copy()
            if fac_sel:
                df_f = df_f[df_f["FACULTAD NOMPROPIO"].isin(fac_sel)]
            if busqueda:
                df_f = df_f[df_f["NOMBRE CARRERA FINAL"].str.contains(busqueda, case=False, na=False)]
            years_sel = [y for y in year_all if rango[0] <= y <= rango[1]]

            st.markdown(f"""
            <div class="table-wrapper">
              <div class="table-title-bar">
                ADMISIÓN TOTAL ACUMULADA AL 30 DE ABRIL (1ER SEM y 1ER TRIM)
                &nbsp;·&nbsp; {len(df_f)} carreras · {int(years_sel[0])}–{int(years_sel[-1])}
              </div>
              <div class="table-scroll">{build_t1(df_f, years_sel, heatmap)}</div>
            </div>""", unsafe_allow_html=True)

            exp_cols = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"] + years_sel
            df_exp   = df_f[[c for c in exp_cols if c in df_f.columns]].copy()
            df_exp.columns = ["FACULTAD", "CÓDIGO SIGA", "CARRERA"] + [int(y) for y in years_sel]
            st.session_state["export_adm_carrera"] = df_exp
            st.session_state["export_adm_carrera_raw"] = (df_f, years_sel)

            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                df_exp.to_excel(w, index=False, sheet_name="Admisión por Carrera")

            st.markdown(
                dl_link(buf.getvalue(), "UAH_admision_carrera.xlsx",
                        "⬇  Descargar Excel",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                + dl_link(df_exp.to_csv(index=False).encode("utf-8-sig"),
                          "UAH_admision_carrera.csv", "⬇  Descargar CSV", "text/csv"),
                unsafe_allow_html=True
            )

        with tab2:
            try:
                with st.spinner("Procesando datos de sexo…"):
                    df_sex = procesar_sexo(file_bytes, hoja_sel)
            except Exception as e:
                st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

            year_sex = sorted(df_sex["AÑO INGRESO"].unique())
            yr_s     = year_sex[-1]
            df_s_yr  = df_sex[df_sex["AÑO INGRESO"] == yr_s]
            n_muj    = int(df_s_yr[df_s_yr["SEXO FINAL"] == "M"]["N"].sum())
            n_hom    = int(df_s_yr[df_s_yr["SEXO FINAL"] == "H"]["N"].sum())
            n_tot    = n_muj + n_hom

            st.markdown(f"""
            <div class="metric-strip">
              <div class="metric-card"><div class="mc-label">Carreras</div>
                <div class="mc-value">{int(df_sex[(df_sex["AÑO INGRESO"] == yr_s) & (df_sex["N"] > 0)]["CODIGO SIGA"].nunique())}</div></div>
              <div class="metric-card"><div class="mc-label">Mujeres {int(yr_s)}</div>
                <div class="mc-value">{n_muj:,}</div></div>
              <div class="metric-card"><div class="mc-label">Hombres {int(yr_s)}</div>
                <div class="mc-value">{n_hom:,}</div></div>
              <div class="metric-card accent"><div class="mc-label">% Mujeres {int(yr_s)}</div>
                <div class="mc-value">{n_muj/n_tot*100:.0f}%</div></div>
              <div class="metric-card"><div class="mc-label">Años</div>
                <div class="mc-value">{int(year_sex[0])}–{int(year_sex[-1])}</div></div>
            </div>""", unsafe_allow_html=True)

            g1, g2, g3 = st.columns([2, 2, 2])
            with g1:
                facs_s  = sorted(df_sex["FACULTAD NOMPROPIO"].dropna().unique())
                fac_s   = st.multiselect("Facultad", facs_s, default=facs_s,
                                          placeholder="Todas", key="fac_sex")
            with g2:
                rango_s = st.select_slider("Rango de años",
                                           options=[int(y) for y in year_sex],
                                           value=(int(year_sex[0]), int(year_sex[-1])),
                                           key="rango_sex")
            with g3:
                busq_s  = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…",
                                        key="busq_sex")

            df_sf = df_sex.copy()
            if fac_s:
                df_sf = df_sf[df_sf["FACULTAD NOMPROPIO"].isin(fac_s)]
            if busq_s:
                df_sf = df_sf[df_sf["NOMBRE CARRERA FINAL"].str.contains(busq_s, case=False, na=False)]
            years_s_sel = [y for y in year_sex if rango_s[0] <= y <= rango_s[1]]
            df_sf = df_sf[df_sf["AÑO INGRESO"].isin(years_s_sel)]

            st.markdown(f"""
            <div class="table-wrapper">
              <div class="table-title-bar">
                ADMISIÓN PREGRADO SEGÚN SEXO — NÚMERO, PORCENTAJE Y TOTAL
                &nbsp;·&nbsp; {df_sf["CODIGO SIGA"].nunique()} carreras · {int(years_s_sel[0])}–{int(years_s_sel[-1])}
              </div>
              <div class="table-scroll">{build_t2(df_sf, years_s_sel)}</div>
            </div>""", unsafe_allow_html=True)

            df_dl = df_sf[["FACULTAD NOMPROPIO", "JORNADA", "CODIGO SIGA",
                           "NOMBRE CARRERA FINAL", "AÑO INGRESO", "SEXO FINAL",
                           "N", "PCT", "Total"]].copy()
            df_dl.columns = ["FACULTAD", "JORNADA", "CÓD.SIGA", "CARRERA",
                             "AÑO", "SEXO", "N", "PORCENTAJE", "TOTAL"]
            st.session_state["export_adm_sexo"] = df_dl
            st.session_state["export_adm_sexo_raw"] = (df_sf, years_s_sel)

            buf2 = io.BytesIO()
            with pd.ExcelWriter(buf2, engine="openpyxl") as w:
                df_dl.to_excel(w, index=False, sheet_name="Admisión por Sexo")

            st.markdown(
                dl_link(buf2.getvalue(), "UAH_admision_sexo.xlsx",
                        "⬇  Descargar Excel",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                + dl_link(df_dl.to_csv(index=False).encode("utf-8-sig"),
                          "UAH_admision_sexo.csv", "⬇  Descargar CSV", "text/csv"),
                unsafe_allow_html=True
            )

        # ══════════════════════════════════════════════════════════════════
        # SUB-TAB 3: ADMISIÓN POR VÍA DE INGRESO
        # ══════════════════════════════════════════════════════════════════
        with tab3:
            try:
                with st.spinner("Procesando…"):
                    df_via = procesar_via_ingreso(file_bytes, hoja_sel)
            except Exception as e:
                st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

            year_via = sorted(df_via["AÑO INGRESO"].unique())
            yr_via   = year_via[-1]
            df_via_yr = df_via[df_via["AÑO INGRESO"] == yr_via]
            tot_via  = int(df_via_yr.drop_duplicates(subset=["CODIGO SIGA","AÑO INGRESO"])["Total"].sum())
            ord_via  = int(df_via_yr[df_via_yr["VIA"] == "Adm. Ordinaria DEMRE"]["N"].sum())
            pace_via = int(df_via_yr[df_via_yr["VIA"] == "PACE"]["N"].sum())
            esp_via  = int(df_via_yr[df_via_yr["VIA"] == "Adm. Especial (*)"]["N"].sum())

            st.markdown(f"""
            <div class="metric-strip">
              <div class="metric-card"><div class="mc-label">Total {int(yr_via)}</div>
                <div class="mc-value">{tot_via:,}</div></div>
              <div class="metric-card"><div class="mc-label">Ordinaria {int(yr_via)}</div>
                <div class="mc-value">{ord_via:,}</div></div>
              <div class="metric-card"><div class="mc-label">Adm. Especial {int(yr_via)}</div>
                <div class="mc-value">{esp_via:,}</div></div>
              <div class="metric-card accent"><div class="mc-label">PACE {int(yr_via)}</div>
                <div class="mc-value">{pace_via:,}</div></div>
              <div class="metric-card"><div class="mc-label">Años</div>
                <div class="mc-value">{int(year_via[0])}–{int(year_via[-1])}</div></div>
            </div>""", unsafe_allow_html=True)

            fv1, fv2, fv3 = st.columns([2, 2, 2])
            with fv1:
                facs_via  = sorted(df_via["FACULTAD NOMPROPIO"].dropna().unique())
                fac_via   = st.multiselect("Facultad", facs_via, default=facs_via,
                                           placeholder="Todas", key="fac_via")
            with fv2:
                rango_via = st.select_slider("Rango de años",
                                             options=[int(y) for y in year_via],
                                             value=(int(year_via[0]), int(year_via[-1])),
                                             key="rango_via")
            with fv3:
                busq_via  = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…",
                                          key="busq_via")

            df_fv = df_via.copy()
            if fac_via:
                df_fv = df_fv[df_fv["FACULTAD NOMPROPIO"].isin(fac_via)]
            if busq_via:
                df_fv = df_fv[df_fv["NOMBRE CARRERA FINAL"].str.contains(busq_via, case=False, na=False)]
            years_via_sel = [y for y in year_via if rango_via[0] <= y <= rango_via[1]]
            df_fv = df_fv[df_fv["AÑO INGRESO"].isin(years_via_sel)]

            st.markdown(f"""
            <div class="table-wrapper">
              <div class="table-title-bar">
                ADMISIÓN POR VÍA DE INGRESO
                &nbsp;·&nbsp; {df_fv["CODIGO SIGA"].nunique()} carreras · {int(years_via_sel[0])}–{int(years_via_sel[-1])}
              </div>
              <div class="table-scroll">{build_t5(df_fv, years_via_sel)}</div>
            </div>""", unsafe_allow_html=True)

            st.session_state["export_adm_via_raw"] = (df_fv, years_via_sel)

            buf_via = io.BytesIO()
            import openpyxl as _xl
            _wb = _xl.Workbook(); _wb.remove(_wb.active)
            _ws = _wb.create_sheet("Adm por Vía Ingreso")
            excel_t5(_ws, df_fv, years_via_sel, "ADMISIÓN POR VÍA DE INGRESO")
            _wb.save(buf_via)
            st.markdown(
                dl_link(buf_via.getvalue(), "UAH_admision_via_ingreso.xlsx",
                        "⬇  Descargar Excel",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                + dl_link(df_fv.to_csv(index=False).encode("utf-8-sig"),
                          "UAH_admision_via_ingreso.csv", "⬇  Descargar CSV", "text/csv"),
                unsafe_allow_html=True
            )

        # ══════════════════════════════════════════════════════════════════
        # SUB-TAB 4: ADMISIÓN REGULAR
        # ══════════════════════════════════════════════════════════════════
        with tab4:
            try:
                with st.spinner("Procesando admisión regular…"):
                    df_reg = procesar_admision_regular(file_bytes, hoja_sel)
            except Exception as e:
                st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

            year_all_r = sorted([c for c in df_reg.columns
                                  if isinstance(c, (int, float)) and 2000 <= float(c) <= 2030])
            if not year_all_r:
                st.warning("No se encontraron datos con indicador '1.OK'."); st.stop()

            yr_last_r      = year_all_r[-1]
            total_ultimo_r = int(df_reg[yr_last_r].sum()) if yr_last_r in df_reg.columns else 0
            total_hist_r   = int(sum(df_reg[y].sum() for y in year_all_r if y in df_reg.columns))

            st.markdown(f"""
            <div class="metric-strip">
              <div class="metric-card"><div class="mc-label">Carreras</div>
                <div class="mc-value">{int((df_reg[yr_last_r] > 0).sum())}</div></div>
              <div class="metric-card"><div class="mc-label">Facultades</div>
                <div class="mc-value">{df_reg["FACULTAD NOMPROPIO"].nunique()}</div></div>
              <div class="metric-card accent"><div class="mc-label">Admitidos {int(yr_last_r)}</div>
                <div class="mc-value">{total_ultimo_r:,}</div></div>
              <div class="metric-card"><div class="mc-label">Total histórico</div>
                <div class="mc-value">{total_hist_r:,}</div></div>
              <div class="metric-card"><div class="mc-label">Años</div>
                <div class="mc-value">{int(year_all_r[0])}–{int(year_all_r[-1])}</div></div>
            </div>""", unsafe_allow_html=True)

            fr1, fr2, fr3, fr4 = st.columns([2, 2, 2, 1])
            with fr1:
                facs_r    = sorted(df_reg["FACULTAD NOMPROPIO"].dropna().unique())
                fac_sel_r = st.multiselect("Facultad", facs_r, default=facs_r,
                                           placeholder="Todas", key="fac_reg")
            with fr2:
                rango_r = st.select_slider("Rango de años",
                                           options=[int(y) for y in year_all_r],
                                           value=(int(year_all_r[0]), int(year_all_r[-1])),
                                           key="rango_reg")
            with fr3:
                busqueda_r = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…",
                                           key="busq_reg")
            with fr4:
                heatmap_r = st.toggle("Mapa de calor", value=True, key="hm_reg")

            df_fr = df_reg.copy()
            if fac_sel_r:
                df_fr = df_fr[df_fr["FACULTAD NOMPROPIO"].isin(fac_sel_r)]
            if busqueda_r:
                df_fr = df_fr[df_fr["NOMBRE CARRERA FINAL"].str.contains(busqueda_r, case=False, na=False)]
            years_sel_r = [y for y in year_all_r if rango_r[0] <= y <= rango_r[1]]

            st.markdown(f"""
            <div class="table-wrapper">
              <div class="table-title-bar">
                ADMISIÓN REGULAR (Indicador: 1.OK)
                &nbsp;·&nbsp; {len(df_fr)} carreras · {int(years_sel_r[0])}–{int(years_sel_r[-1])}
              </div>
              <div class="table-scroll">{build_t1(df_fr, years_sel_r, heatmap_r)}</div>
            </div>""", unsafe_allow_html=True)

            exp_cols_r = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"] + years_sel_r
            df_exp_r   = df_fr[[c for c in exp_cols_r if c in df_fr.columns]].copy()
            df_exp_r.columns = ["FACULTAD", "CÓDIGO SIGA", "CARRERA"] + [int(y) for y in years_sel_r]
            st.session_state["export_adm_regular"] = df_exp_r
            st.session_state["export_adm_regular_raw"] = (df_fr, years_sel_r)

            buf_r = io.BytesIO()
            with pd.ExcelWriter(buf_r, engine="openpyxl") as w:
                df_exp_r.to_excel(w, index=False, sheet_name="Admisión Regular")

            st.markdown(
                dl_link(buf_r.getvalue(), "UAH_admision_regular.xlsx",
                        "⬇  Descargar Excel",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                + dl_link(df_exp_r.to_csv(index=False).encode("utf-8-sig"),
                          "UAH_admision_regular.csv", "⬇  Descargar CSV", "text/csv"),
                unsafe_allow_html=True
            )

        # ── Exportación global Admisión ──
        sheets_adm = {
            "Adm por Carrera": st.session_state.get("export_adm_carrera"),
            "Adm por Sexo":    st.session_state.get("export_adm_sexo"),
            "Adm por Vía":     st.session_state.get("export_adm_via_raw"),
            "Adm Regular":     st.session_state.get("export_adm_regular"),
        }
        if all(v is not None for v in sheets_adm.values()):
            buf_adm_all = io.BytesIO()
            with pd.ExcelWriter(buf_adm_all, engine="openpyxl") as w:
                for sheet_name, df_s in sheets_adm.items():
                    if isinstance(df_s, tuple):
                        df_s[0].to_excel(w, index=False, sheet_name=sheet_name)
                    else:
                        df_s.to_excel(w, index=False, sheet_name=sheet_name)
            st.markdown("---")
            st.markdown(
                dl_link(buf_adm_all.getvalue(), "UAH_admision_completo.xlsx",
                        "⬇  Exportar todas las tablas de Admisión (Excel)",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                unsafe_allow_html=True
            )


# ══════════════════════════════════════════════════════════════════════════════
# PESTAÑA PRINCIPAL: MATRÍCULA COMPLETA
# ══════════════════════════════════════════════════════════════════════════════
with main_mat:

    file_bytes_mat, hoja_sel_mat = selector_origen_archivo(
        nombre="Matrícula Completa",
        ejemplo="01. MATRÍCULA TOTAL ACUMULADA 2016-2026.xlsx",
        key_prefix="mat",
        archivo_enc="matricula.enc",
    )

    if not file_bytes_mat:
        st.info("⬆️  Sube un archivo Excel de Matrícula para comenzar.")
    else:
        st.markdown(f'<div class="sheet-badge">🗂 &nbsp;Hoja activa: <strong>{hoja_sel_mat}</strong></div>',
                    unsafe_allow_html=True)

        tab_m1, tab_m2, tab_m3, tab_m4, tab_m5 = st.tabs([
            "📊  Matrícula por Carrera",
            "👥  Matrícula por Sexo",
            "🎓  Nueva Matrícula y Cursos Superiores",
            "🔁  Retención 1er Año",
            "📈  Seguimiento por Carrera y Cohorte",
        ])

        with tab_m1:
            try:
                with st.spinner("Procesando…"):
                    df_mat = procesar_matricula(file_bytes_mat, hoja_sel_mat)
            except Exception as e:
                st.error(f"No se pudo leer la hoja.\n\n`{e}`"); st.stop()

            year_all_m = sorted([c for c in df_mat.columns
                                 if isinstance(c, (int, float)) and 2000 <= float(c) <= 2030])
            if not year_all_m:
                st.warning("La hoja no contiene columnas de año."); st.stop()

            yr_last_m      = year_all_m[-1]
            _conteo_m      = contar_matriculados_por_anio(file_bytes_mat, hoja_sel_mat)
            total_ultimo_m = _conteo_m.get(yr_last_m, 0)
            total_hist_m   = sum(_conteo_m.values())

            st.markdown(f"""
            <div class="metric-strip">
              <div class="metric-card"><div class="mc-label">Carreras</div>
                <div class="mc-value">{int((df_mat[yr_last_m] > 0).sum())}</div></div>
              <div class="metric-card"><div class="mc-label">Facultades</div>
                <div class="mc-value">{df_mat["FACULTAD NOMPROPIO"].nunique()}</div></div>
              <div class="metric-card"><div class="mc-label">Matriculados {int(yr_last_m)}</div>
                <div class="mc-value">{total_ultimo_m:,}</div></div>
              <div class="metric-card accent"><div class="mc-label">Total histórico</div>
                <div class="mc-value">{total_hist_m:,}</div></div>
              <div class="metric-card"><div class="mc-label">Años</div>
                <div class="mc-value">{int(year_all_m[0])}–{int(year_all_m[-1])}</div></div>
            </div>""", unsafe_allow_html=True)

            fm1, fm2, fm3, fm4 = st.columns([2, 2, 2, 1])
            with fm1:
                facs_m    = sorted(df_mat["FACULTAD NOMPROPIO"].dropna().unique())
                fac_sel_m = st.multiselect("Facultad", facs_m, default=facs_m,
                                           placeholder="Todas", key="fac_mat")
            with fm2:
                rango_m = st.select_slider("Rango de años",
                                           options=[int(y) for y in year_all_m],
                                           value=(int(year_all_m[0]), int(year_all_m[-1])),
                                           key="rango_mat")
            with fm3:
                busqueda_m = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…",
                                           key="busq_mat")
            with fm4:
                heatmap_m = st.toggle("Mapa de calor", value=True, key="hm_mat")

            df_fm = df_mat.copy()
            if fac_sel_m:
                df_fm = df_fm[df_fm["FACULTAD NOMPROPIO"].isin(fac_sel_m)]
            if busqueda_m:
                df_fm = df_fm[df_fm["NOMBRE CARRERA FINAL"].str.contains(busqueda_m, case=False, na=False)]
            years_sel_m = [y for y in year_all_m if rango_m[0] <= y <= rango_m[1]]

            st.markdown(f"""
            <div class="table-wrapper">
              <div class="table-title-bar">
                MATRÍCULA TOTAL ACUMULADA
                &nbsp;·&nbsp; {len(df_fm)} carreras · {int(years_sel_m[0])}–{int(years_sel_m[-1])}
              </div>
              <div class="table-scroll">{build_t1(df_fm, years_sel_m, heatmap_m, show_diff=True)}</div>
            </div>""", unsafe_allow_html=True)

            exp_cols_m = ["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"] + years_sel_m
            df_exp_m   = df_fm[[c for c in exp_cols_m if c in df_fm.columns]].copy()
            df_exp_m.columns = ["FACULTAD", "CÓDIGO SIGA", "CARRERA"] + [int(y) for y in years_sel_m]
            st.session_state["export_mat_carrera"] = df_exp_m
            st.session_state["export_mat_carrera_raw"] = (df_fm, years_sel_m)

            buf_m = io.BytesIO()
            with pd.ExcelWriter(buf_m, engine="openpyxl") as w:
                df_exp_m.to_excel(w, index=False, sheet_name="Matrícula por Carrera")

            st.markdown(
                dl_link(buf_m.getvalue(), "UAH_matricula_carrera.xlsx",
                        "⬇  Descargar Excel",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                + dl_link(df_exp_m.to_csv(index=False).encode("utf-8-sig"),
                          "UAH_matricula_carrera.csv", "⬇  Descargar CSV", "text/csv"),
                unsafe_allow_html=True
            )

        with tab_m2:
            try:
                with st.spinner("Procesando datos de sexo…"):
                    df_sex_m = procesar_matricula_sexo(file_bytes_mat, hoja_sel_mat)
            except Exception as e:
                st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

            year_sex_m = sorted(df_sex_m["AÑO INGRESO"].unique())
            yr_sm      = year_sex_m[-1]
            df_s_yr_m  = df_sex_m[df_sex_m["AÑO INGRESO"] == yr_sm]
            n_muj_m    = int(df_s_yr_m[df_s_yr_m["SEXO FINAL"] == "M"]["N"].sum())
            n_hom_m    = int(df_s_yr_m[df_s_yr_m["SEXO FINAL"] == "H"]["N"].sum())
            n_tot_m    = n_muj_m + n_hom_m

            st.markdown(f"""
            <div class="metric-strip">
              <div class="metric-card"><div class="mc-label">Carreras</div>
                <div class="mc-value">{int(df_sex_m[(df_sex_m["AÑO INGRESO"] == yr_sm) & (df_sex_m["N"] > 0)]["CODIGO SIGA"].nunique())}</div></div>
              <div class="metric-card"><div class="mc-label">Mujeres {int(yr_sm)}</div>
                <div class="mc-value">{n_muj_m:,}</div></div>
              <div class="metric-card"><div class="mc-label">Hombres {int(yr_sm)}</div>
                <div class="mc-value">{n_hom_m:,}</div></div>
              <div class="metric-card accent"><div class="mc-label">% Mujeres {int(yr_sm)}</div>
                <div class="mc-value">{n_muj_m/n_tot_m*100:.0f}%</div></div>
              <div class="metric-card"><div class="mc-label">Años</div>
                <div class="mc-value">{int(year_sex_m[0])}–{int(year_sex_m[-1])}</div></div>
            </div>""", unsafe_allow_html=True)

            gm1, gm2, gm3 = st.columns([2, 2, 2])
            with gm1:
                facs_sm  = sorted(df_sex_m["FACULTAD NOMPROPIO"].dropna().unique())
                fac_sm   = st.multiselect("Facultad", facs_sm, default=facs_sm,
                                          placeholder="Todas", key="fac_sex_mat")
            with gm2:
                rango_sm = st.select_slider("Rango de años",
                                            options=[int(y) for y in year_sex_m],
                                            value=(int(year_sex_m[0]), int(year_sex_m[-1])),
                                            key="rango_sex_mat")
            with gm3:
                busq_sm  = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…",
                                         key="busq_sex_mat")

            df_sfm = df_sex_m.copy()
            if fac_sm:
                df_sfm = df_sfm[df_sfm["FACULTAD NOMPROPIO"].isin(fac_sm)]
            if busq_sm:
                df_sfm = df_sfm[df_sfm["NOMBRE CARRERA FINAL"].str.contains(busq_sm, case=False, na=False)]
            years_sm_sel = [y for y in year_sex_m if rango_sm[0] <= y <= rango_sm[1]]
            df_sfm = df_sfm[df_sfm["AÑO INGRESO"].isin(years_sm_sel)]

            st.markdown(f"""
            <div class="table-wrapper">
              <div class="table-title-bar">
                MATRÍCULA SEGÚN SEXO — NÚMERO, PORCENTAJE Y TOTAL
                &nbsp;·&nbsp; {df_sfm["CODIGO SIGA"].nunique()} carreras · {int(years_sm_sel[0])}–{int(years_sm_sel[-1])}
              </div>
              <div class="table-scroll">{build_t2(df_sfm, years_sm_sel)}</div>
            </div>""", unsafe_allow_html=True)

            df_dl_m = df_sfm[["FACULTAD NOMPROPIO", "JORNADA", "CODIGO SIGA",
                               "NOMBRE CARRERA FINAL", "AÑO INGRESO", "SEXO FINAL",
                               "N", "PCT", "Total"]].copy()
            df_dl_m.columns = ["FACULTAD", "JORNADA", "CÓD.SIGA", "CARRERA",
                               "AÑO", "SEXO", "N", "PORCENTAJE", "TOTAL"]
            st.session_state["export_mat_sexo"] = df_dl_m
            st.session_state["export_mat_sexo_raw"] = (df_sfm, years_sm_sel)
            buf_m2 = io.BytesIO()
            with pd.ExcelWriter(buf_m2, engine="openpyxl") as w:
                df_dl_m.to_excel(w, index=False, sheet_name="Matrícula por Sexo")

            st.markdown(
                dl_link(buf_m2.getvalue(), "UAH_matricula_sexo.xlsx",
                        "⬇  Descargar Excel",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                + dl_link(df_dl_m.to_csv(index=False).encode("utf-8-sig"),
                          "UAH_matricula_sexo.csv", "⬇  Descargar CSV", "text/csv"),
                unsafe_allow_html=True
            )

        # ══════════════════════════════════════════════════════════════════
        # SUB-TAB 3: NUEVA MATRÍCULA Y CURSOS SUPERIORES
        # ══════════════════════════════════════════════════════════════════
        with tab_m3:
            try:
                with st.spinner("Procesando…"):
                    df_ncs = procesar_matricula_nueva(file_bytes_mat, hoja_sel_mat)
            except Exception as e:
                st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

            year_ncs  = sorted(df_ncs["AÑO"].unique())
            yr_ncs    = year_ncs[-1]
            df_ncs_yr = df_ncs[df_ncs["AÑO"] == yr_ncs]
            n_nuevo   = int(df_ncs_yr[df_ncs_yr["TIPO"] == "NUEVO"]["N"].sum())
            n_sup     = int(df_ncs_yr[df_ncs_yr["TIPO"] == "SUPERIOR"]["N"].sum())
            n_tot_ncs = n_nuevo + n_sup

            st.markdown(f"""
            <div class="metric-strip">
              <div class="metric-card"><div class="mc-label">Nuevos {int(yr_ncs)}</div>
                <div class="mc-value">{n_nuevo:,}</div></div>
              <div class="metric-card"><div class="mc-label">Cursos Sup. {int(yr_ncs)}</div>
                <div class="mc-value">{n_sup:,}</div></div>
              <div class="metric-card accent"><div class="mc-label">% Nuevos {int(yr_ncs)}</div>
                <div class="mc-value">{n_nuevo/n_tot_ncs*100:.0f}%</div></div>
              <div class="metric-card"><div class="mc-label">Total {int(yr_ncs)}</div>
                <div class="mc-value">{n_tot_ncs:,}</div></div>
              <div class="metric-card"><div class="mc-label">Años</div>
                <div class="mc-value">{int(year_ncs[0])}–{int(year_ncs[-1])}</div></div>
            </div>""", unsafe_allow_html=True)

            hn1, hn2, hn3 = st.columns([2, 2, 2])
            with hn1:
                facs_ncs  = sorted(df_ncs["FACULTAD NOMPROPIO"].dropna().unique())
                fac_ncs   = st.multiselect("Facultad", facs_ncs, default=facs_ncs,
                                           placeholder="Todas", key="fac_ncs")
            with hn2:
                rango_ncs = st.select_slider("Rango de años",
                                             options=[int(y) for y in year_ncs],
                                             value=(int(year_ncs[0]), int(year_ncs[-1])),
                                             key="rango_ncs")
            with hn3:
                busq_ncs  = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…",
                                          key="busq_ncs")

            df_fn = df_ncs.copy()
            if fac_ncs:
                df_fn = df_fn[df_fn["FACULTAD NOMPROPIO"].isin(fac_ncs)]
            if busq_ncs:
                df_fn = df_fn[df_fn["NOMBRE CARRERA FINAL"].str.contains(busq_ncs, case=False, na=False)]
            years_ncs_sel = [y for y in year_ncs if rango_ncs[0] <= y <= rango_ncs[1]]
            df_fn = df_fn[df_fn["AÑO"].isin(years_ncs_sel)]

            st.markdown(f"""
            <div class="table-wrapper">
              <div class="table-title-bar">
                NUEVA MATRÍCULA Y CURSOS SUPERIORES
                &nbsp;·&nbsp; {df_fn["CODIGO SIGA"].nunique()} carreras · {int(years_ncs_sel[0])}–{int(years_ncs_sel[-1])}
              </div>
              <div class="table-scroll">{build_t3(df_fn, years_ncs_sel)}</div>
            </div>""", unsafe_allow_html=True)

            df_dl_ncs = df_fn[["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL",
                                "AÑO", "TIPO", "N", "PCT", "Total"]].copy()
            df_dl_ncs.columns = ["FACULTAD", "CÓD.SIGA", "CARRERA", "AÑO", "TIPO", "N", "PORCENTAJE", "TOTAL"]
            st.session_state["export_mat_nueva"] = df_dl_ncs
            st.session_state["export_mat_nueva_raw"] = (df_fn, years_ncs_sel)

            buf_ncs = io.BytesIO()
            with pd.ExcelWriter(buf_ncs, engine="openpyxl") as w:
                df_dl_ncs.to_excel(w, index=False, sheet_name="Nueva Mat y Cursos Sup")

            st.markdown(
                dl_link(buf_ncs.getvalue(), "UAH_matricula_nueva_sup.xlsx",
                        "⬇  Descargar Excel",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                + dl_link(df_dl_ncs.to_csv(index=False).encode("utf-8-sig"),
                          "UAH_matricula_nueva_sup.csv", "⬇  Descargar CSV", "text/csv"),
                unsafe_allow_html=True
            )

        # ══════════════════════════════════════════════════════════════════
        # SUB-TAB 4: RETENCIÓN 1ER AÑO
        # ══════════════════════════════════════════════════════════════════
        with tab_m4:
            if not file_bytes:
                st.warning("⚠️  Para calcular retención también debes cargar el archivo de **Admisión** (pestaña 🎓 Admisión).")
            else:
                try:
                    with st.spinner("Calculando retención…"):
                        df_ret = procesar_retencion(
                            file_bytes, hoja_sel,
                            file_bytes_mat, hoja_sel_mat
                        )
                except Exception as e:
                    st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

                cohortes_all = sorted(df_ret["AÑO COHORTE"].unique())
                yr_ret    = cohortes_all[-1]
                ret_yr    = df_ret[df_ret["AÑO COHORTE"] == yr_ret]
                ret_uah   = (ret_yr["N_RETENIDOS"].sum() / ret_yr["N_COHORTE"].sum() * 100
                             if ret_yr["N_COHORTE"].sum() > 0 else 0)
                ret_prev  = cohortes_all[-2] if len(cohortes_all) >= 2 else None
                ret_prev_pct = 0
                if ret_prev:
                    ret_prev_val = df_ret[df_ret["AÑO COHORTE"] == ret_prev]
                    ret_prev_pct = (ret_prev_val["N_RETENIDOS"].sum() /
                                    ret_prev_val["N_COHORTE"].sum() * 100
                                    if ret_prev_val["N_COHORTE"].sum() > 0 else 0)

                st.markdown(f"""
                <div class="metric-strip">
                  <div class="metric-card"><div class="mc-label">Carreras</div>
                    <div class="mc-value">{df_ret["CODIGO SIGA"].nunique()}</div></div>
                  <div class="metric-card accent"><div class="mc-label">Retención UAH Coh. {int(yr_ret)}</div>
                    <div class="mc-value">{ret_uah:.0f}%</div></div>
                  {"" if not ret_prev else f'<div class="metric-card"><div class="mc-label">Retención UAH Coh. {int(ret_prev)}</div><div class="mc-value">{ret_prev_pct:.0f}%</div></div>'}
                  <div class="metric-card"><div class="mc-label">Cohortes</div>
                    <div class="mc-value">{int(cohortes_all[0])}–{int(cohortes_all[-1])}</div></div>
                </div>""", unsafe_allow_html=True)

                hr1, hr2, hr3 = st.columns([2, 2, 2])
                with hr1:
                    facs_ret  = sorted(df_ret["FACULTAD NOMPROPIO"].dropna().unique())
                    fac_ret   = st.multiselect("Facultad", facs_ret, default=facs_ret,
                                               placeholder="Todas", key="fac_ret")
                with hr2:
                    rango_ret = st.select_slider("Rango de cohortes",
                                                 options=[int(y) for y in cohortes_all],
                                                 value=(int(cohortes_all[0]), int(cohortes_all[-1])),
                                                 key="rango_ret")
                with hr3:
                    busq_ret  = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…",
                                              key="busq_ret")

                df_fr = df_ret.copy()
                if fac_ret:
                    df_fr = df_fr[df_fr["FACULTAD NOMPROPIO"].isin(fac_ret)]
                if busq_ret:
                    df_fr = df_fr[df_fr["NOMBRE CARRERA FINAL"].str.contains(busq_ret, case=False, na=False)]
                cohortes_sel = [y for y in cohortes_all if rango_ret[0] <= y <= rango_ret[1]]
                df_fr = df_fr[df_fr["AÑO COHORTE"].isin(cohortes_sel)]

                st.markdown(f"""
                <div class="table-wrapper">
                  <div class="table-title-bar">
                    RETENCIÓN DE 1ER AÑO (30 de Abril del año siguiente)
                    &nbsp;·&nbsp; {df_fr["CODIGO SIGA"].nunique()} carreras · Coh. {int(cohortes_sel[0])}–{int(cohortes_sel[-1])}
                  </div>
                  <div class="table-scroll">{build_t4(df_fr, cohortes_sel)}</div>
                </div>""", unsafe_allow_html=True)

                pivot_dl = df_fr.pivot_table(
                    index=["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"],
                    columns="AÑO COHORTE", values="PCT", aggfunc="first"
                ).reset_index()
                pivot_dl.columns.name = None
                yr_cols = [c for c in pivot_dl.columns if isinstance(c, (int, float))]
                for c in yr_cols:
                    pivot_dl[c] = pivot_dl[c].apply(lambda v: f"{v*100:.0f}%" if pd.notna(v) else "—")
                pivot_dl.columns = (["FACULTAD", "CÓD.SIGA", "CARRERA"] +
                                     [f"Coh. {int(c)}" for c in yr_cols])
                st.session_state["export_ret_raw"] = (df_fr, cohortes_sel)

                buf_ret = io.BytesIO()
                with pd.ExcelWriter(buf_ret, engine="openpyxl") as w:
                    pivot_dl.to_excel(w, index=False, sheet_name="Retención 1er Año")
                st.markdown(
                    dl_link(buf_ret.getvalue(), "UAH_retencion_1er_anio.xlsx",
                            "⬇  Descargar Excel",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    + dl_link(pivot_dl.to_csv(index=False).encode("utf-8-sig"),
                              "UAH_retencion_1er_anio.csv", "⬇  Descargar CSV", "text/csv"),
                    unsafe_allow_html=True
                )

        # ══════════════════════════════════════════════════════════════════
        # SUB-TAB 5: SEGUIMIENTO POR CARRERA Y COHORTE
        # ══════════════════════════════════════════════════════════════════
        with tab_m5:
            if not file_bytes:
                st.warning("⚠️  Para el seguimiento de cohortes también debes cargar el archivo de **Admisión** (pestaña 🎓 Admisión).")
            else:
                try:
                    with st.spinner("Calculando seguimiento de cohortes…"):
                        df_seg = procesar_seguimiento_cohorte(
                            file_bytes, hoja_sel,
                            file_bytes_mat, hoja_sel_mat
                        )
                except Exception as e:
                    st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

                import plotly.graph_objects as go

                cohortes_seg_all = sorted(df_seg["AÑO COHORTE"].unique())
                carreras_seg_all = sorted(df_seg["NOMBRE CARRERA FINAL"].dropna().unique())
                facs_seg_all     = sorted(df_seg["FACULTAD NOMPROPIO"].dropna().unique())

                # ── Métricas resumen ──
                n_cohortes_seg = len(cohortes_seg_all)
                n_carreras_seg = df_seg["CODIGO SIGA"].nunique()
                st.markdown(f"""
                <div class="metric-strip">
                  <div class="metric-card"><div class="mc-label">Carreras con seguimiento</div>
                    <div class="mc-value">{n_carreras_seg}</div></div>
                  <div class="metric-card"><div class="mc-label">Cohortes disponibles</div>
                    <div class="mc-value">{n_cohortes_seg}</div></div>
                  <div class="metric-card"><div class="mc-label">Período</div>
                    <div class="mc-value">{int(cohortes_seg_all[0])}–{int(cohortes_seg_all[-1])}</div></div>
                </div>""", unsafe_allow_html=True)

                # ── Filtros ──
                sf1, sf2, sf3 = st.columns([2, 2, 2])
                with sf1:
                    fac_seg = st.selectbox(
                        "Facultad",
                        ["Todas"] + facs_seg_all,
                        key="fac_seg_sel"
                    )
                with sf2:
                    # Filtrar carreras según facultad seleccionada
                    if fac_seg == "Todas":
                        df_seg_fac = df_seg.copy()
                    else:
                        df_seg_fac = df_seg[df_seg["FACULTAD NOMPROPIO"] == fac_seg].copy()
                    carreras_filt = sorted(df_seg_fac["NOMBRE CARRERA FINAL"].dropna().unique())
                    carrera_seg = st.selectbox(
                        "Carrera",
                        carreras_filt,
                        key="carrera_seg_sel"
                    )
                with sf3:
                    cohortes_disp = sorted(
                        df_seg[df_seg["NOMBRE CARRERA FINAL"] == carrera_seg]["AÑO COHORTE"].unique()
                    )
                    cohortes_seg_sel = st.multiselect(
                        "Cohortes a comparar",
                        [int(c) for c in cohortes_disp],
                        default=[int(c) for c in cohortes_disp],
                        key="cohortes_seg_sel"
                    )

                if not cohortes_seg_sel:
                    st.info("Selecciona al menos una cohorte para graficar.")
                else:
                    df_plot = df_seg[
                        (df_seg["NOMBRE CARRERA FINAL"] == carrera_seg) &
                        (df_seg["AÑO COHORTE"].isin(cohortes_seg_sel))
                    ].copy()

                    # ── Gráfico Plotly: eje X = año calendario, eje Y = N presentes ──
                    # Paleta de colores UAH
                    PALETA = [
                        "#1F3864", "#2F5496", "#C00000", "#2E75B6", "#843C0C",
                        "#375623", "#7B2C8C", "#BF8F00", "#00617E", "#5B5EA6",
                        "#1A7A3C", "#C55A11", "#833000", "#002060", "#7F7F7F",
                    ]

                    fig = go.Figure()

                    for i, coh in enumerate(sorted(cohortes_seg_sel)):
                        df_c = df_plot[df_plot["AÑO COHORTE"] == coh].sort_values("AÑO MATRICULA")
                        if df_c.empty:
                            continue
                        n_coh = int(df_c["N_COHORTE"].iloc[0]) if "N_COHORTE" in df_c.columns else None
                        label = f"Cohorte {coh}"
                        if n_coh:
                            label += f" (N={n_coh})"
                        color = PALETA[i % len(PALETA)]

                        # Punto inicial: N de la cohorte en el año 0
                        años_cal  = df_c["AÑO MATRICULA"].tolist()
                        n_vals    = df_c["N_PRESENTES"].tolist()

                        # Añadir el punto de inicio (año cohorte, N_COHORTE) si no está
                        if n_coh and coh not in años_cal:
                            años_cal  = [coh] + años_cal
                            n_vals    = [n_coh] + n_vals

                        hover_texts = []
                        for a, n in zip(años_cal, n_vals):
                            anio_rel = a - coh
                            pct = n / n_coh * 100 if n_coh else 0
                            hover_texts.append(
                                f"<b>Cohorte {coh}</b><br>"
                                f"Año calendario: {a}<br>"
                                f"Año desde ingreso: {anio_rel}<br>"
                                f"N presentes: {n}<br>"
                                f"% del total cohorte: {pct:.1f}%"
                            )

                        fig.add_trace(go.Scatter(
                            x=años_cal,
                            y=n_vals,
                            mode="lines+markers",
                            name=label,
                            line=dict(color=color, width=2.5),
                            marker=dict(color=color, size=8, symbol="circle"),
                            hovertemplate="%{customdata}<extra></extra>",
                            customdata=hover_texts,
                        ))

                    fig.update_layout(
                        title=dict(
                            text=f"<b>Seguimiento de Cohortes — {carrera_seg}</b>",
                            font=dict(family="Barlow Condensed, Arial", size=16, color="#1F3864"),
                            x=0,
                        ),
                        xaxis=dict(
                            title="Año calendario",
                            tickmode="linear",
                            dtick=1,
                            tickfont=dict(size=11),
                            gridcolor="#EEF1F5",
                        ),
                        yaxis=dict(
                            title="N estudiantes retenidos",
                            tickfont=dict(size=11),
                            gridcolor="#EEF1F5",
                            rangemode="tozero",
                        ),
                        legend=dict(
                            orientation="v",
                            yanchor="top",
                            y=1,
                            xanchor="left",
                            x=1.02,
                            font=dict(size=11),
                        ),
                        plot_bgcolor="white",
                        paper_bgcolor="white",
                        hovermode="x unified",
                        margin=dict(l=60, r=180, t=60, b=60),
                        height=480,
                    )

                    st.plotly_chart(fig, use_container_width=True)

                    # ── Tabla de datos del gráfico ──
                    st.markdown("""
                    <div class="table-wrapper">
                      <div class="table-title-bar">DATOS DEL SEGUIMIENTO (N POR AÑO)</div>
                      <div class="table-scroll">
                    """, unsafe_allow_html=True)

                    # Construir pivot: filas = cohortes, columnas = años desde ingreso
                    pivot_seg = df_plot[df_plot["AÑO COHORTE"].isin(cohortes_seg_sel)].pivot_table(
                        index=["AÑO COHORTE", "N_COHORTE"],
                        columns="ANIOS_DESDE_INGRESO",
                        values="N_PRESENTES",
                        aggfunc="sum"
                    ).reset_index()
                    pivot_seg.columns.name = None
                    anio_cols = [c for c in pivot_seg.columns if isinstance(c, (int, float)) and c >= 0]

                    # Construir HTML de la tabla
                    h_tbl = ['<table class="t1"><thead><tr>']
                    h_tbl.append('<th class="c-siga">COHORTE</th>')
                    h_tbl.append('<th>N Cohorte</th>')
                    for ac in sorted(anio_cols):
                        lbl = "Año 0\n(inicio)" if ac == 0 else f"Año +{int(ac)}"
                        h_tbl.append(f'<th>{lbl}</th>')
                    h_tbl.append('</tr></thead><tbody>')

                    for _, row in pivot_seg.sort_values("AÑO COHORTE").iterrows():
                        coh   = int(row["AÑO COHORTE"])
                        n_c   = int(row["N_COHORTE"]) if pd.notna(row["N_COHORTE"]) else "—"
                        h_tbl.append('<tr>')
                        h_tbl.append(f'<td class="c-siga">{coh}</td>')
                        h_tbl.append(f'<td class="c-num">{n_c:,}</td>' if isinstance(n_c, int) else '<td class="c-empty">—</td>')
                        for ac in sorted(anio_cols):
                            v = row.get(ac)
                            if pd.isna(v):
                                h_tbl.append('<td class="c-empty">—</td>')
                            else:
                                n_v  = int(v)
                                pct  = n_v / n_c * 100 if isinstance(n_c, int) and n_c > 0 else 0
                                col_txt = ("color:#1A7A3C" if pct >= 80
                                           else ("color:#7A5C00" if pct >= 60
                                                 else "color:#C00000"))
                                h_tbl.append(f'<td class="c-num" style="{col_txt}">{n_v:,}</td>')
                        h_tbl.append('</tr>')
                    h_tbl.append('</tbody></table>')
                    st.markdown("".join(h_tbl) + "</div></div>", unsafe_allow_html=True)

                    # ── Descarga ──
                    dl_cols = ["AÑO COHORTE", "N_COHORTE"] + [f"AÑO +{int(ac)}" for ac in sorted(anio_cols)]
                    pivot_dl_seg = pivot_seg.copy()
                    pivot_dl_seg.columns = (
                        ["AÑO COHORTE", "N_COHORTE"] +
                        [f"AÑO +{int(ac)}" for ac in sorted(anio_cols)]
                    )
                    buf_seg = io.BytesIO()
                    with pd.ExcelWriter(buf_seg, engine="openpyxl") as w:
                        pivot_dl_seg.to_excel(w, index=False, sheet_name="Seguimiento Cohortes")
                    st.markdown(
                        dl_link(buf_seg.getvalue(), f"UAH_seguimiento_{carrera_seg[:30]}.xlsx",
                                "⬇  Descargar Excel",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        + dl_link(pivot_dl_seg.to_csv(index=False).encode("utf-8-sig"),
                                  f"UAH_seguimiento_{carrera_seg[:30]}.csv",
                                  "⬇  Descargar CSV", "text/csv"),
                        unsafe_allow_html=True
                    )

        # ── Exportación global Matrícula ──
        sheets_mat = {
            "Mat por Carrera":       st.session_state.get("export_mat_carrera"),
            "Mat por Sexo":          st.session_state.get("export_mat_sexo"),
            "Nueva Mat y Curso Sup": st.session_state.get("export_mat_nueva"),
        }
        if all(v is not None for v in sheets_mat.values()):
            buf_mat_all = io.BytesIO()
            with pd.ExcelWriter(buf_mat_all, engine="openpyxl") as w:
                for sheet_name, df_s in sheets_mat.items():
                    df_s.to_excel(w, index=False, sheet_name=sheet_name)
            st.markdown("---")
            st.markdown(
                dl_link(buf_mat_all.getvalue(), "UAH_matricula_completo.xlsx",
                        "⬇  Exportar todas las tablas de Matrícula (Excel)",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                unsafe_allow_html=True
            )

# ──────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN GLOBAL — ADMISIÓN + MATRÍCULA
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("---")
all_sheets = {
    "Adm por Carrera":       st.session_state.get("export_adm_carrera"),
    "Adm por Sexo":          st.session_state.get("export_adm_sexo"),
    "Adm por Vía Ingreso":   st.session_state.get("export_adm_via_raw"),
    "Mat por Carrera":       st.session_state.get("export_mat_carrera"),
    "Mat por Sexo":          st.session_state.get("export_mat_sexo"),
    "Nueva Mat y Curso Sup": st.session_state.get("export_mat_nueva"),
    "Retención 1er Año":     st.session_state.get("export_ret_raw"),
    "Oferta Académica":      st.session_state.get("export_oferta"),
}
sheets_ready = {k: v for k, v in all_sheets.items() if v is not None}

if sheets_ready:
    buf_all = io.BytesIO()
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    if st.session_state.get("export_adm_carrera_raw"):
        df_r, yrs = st.session_state["export_adm_carrera_raw"]
        ws_a1 = wb.create_sheet("Adm por Carrera")
        excel_t1(ws_a1, df_r, yrs, "ADMISIÓN TOTAL ACUMULADA AL 30 DE ABRIL")

    if st.session_state.get("export_adm_sexo_raw"):
        df_r, yrs = st.session_state["export_adm_sexo_raw"]
        ws_a2 = wb.create_sheet("Adm por Sexo")
        excel_t2(ws_a2, df_r, yrs, "ADMISIÓN PREGRADO SEGÚN SEXO")

    if st.session_state.get("export_adm_via_raw"):
        df_r, yrs = st.session_state["export_adm_via_raw"]
        ws_a3 = wb.create_sheet("Adm por Vía Ingreso")
        excel_t5(ws_a3, df_r, yrs, "ADMISIÓN POR VÍA DE INGRESO")

    if st.session_state.get("export_mat_carrera_raw"):
        df_r, yrs = st.session_state["export_mat_carrera_raw"]
        ws_m1 = wb.create_sheet("Mat por Carrera")
        excel_t1(ws_m1, df_r, yrs, "MATRÍCULA TOTAL ACUMULADA")

    if st.session_state.get("export_mat_sexo_raw"):
        df_r, yrs = st.session_state["export_mat_sexo_raw"]
        ws_m2 = wb.create_sheet("Mat por Sexo")
        excel_t2(ws_m2, df_r, yrs, "MATRÍCULA SEGÚN SEXO")

    if st.session_state.get("export_mat_nueva_raw"):
        df_r, yrs = st.session_state["export_mat_nueva_raw"]
        ws_m3 = wb.create_sheet("Nueva Mat y Curso Sup")
        excel_t3(ws_m3, df_r, yrs, "NUEVA MATRÍCULA Y CURSOS SUPERIORES")

    if st.session_state.get("export_ret_raw"):
        df_r, yrs = st.session_state["export_ret_raw"]
        ws_r = wb.create_sheet("Retención 1er Año")
        # Escribir tabla de retención formateada
        pivot_r = df_r.pivot_table(
            index=["FACULTAD NOMPROPIO", "CODIGO SIGA", "NOMBRE CARRERA FINAL"],
            columns="AÑO COHORTE", values="PCT", aggfunc="first"
        ).reset_index()
        pivot_r.columns.name = None
        pivot_r = pivot_r.sort_values(["FACULTAD NOMPROPIO", "NOMBRE CARRERA FINAL"]).reset_index(drop=True)
        cols_coh = [c for c in yrs if c in pivot_r.columns]
        last3    = cols_coh[-3:] if len(cols_coh) >= 3 else cols_coh
        n_cols   = 3 + len(cols_coh) + (1 if len(cols_coh) >= 2 else 0) + 1
        ws_r.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws_r.cell(1, 1, "RETENCIÓN DE 1ER AÑO (30 de Abril del año siguiente)").fill = _fill(AZ)
        ws_r.cell(1, 1).font = _font(bold=True, size=10)
        ws_r.cell(1, 1).alignment = _align("center")
        hdrs = ["FACULTAD", "CÓDIGO", "CARRERA"] + [f"Coh. {int(c)}" for c in cols_coh]
        if len(cols_coh) >= 2:
            hdrs.append(f"Dif {int(cols_coh[-2])}–{int(cols_coh[-1])}")
        hdrs.append("Prom. últ. 3")
        for ci, h in enumerate(hdrs, 1):
            c = ws_r.cell(2, ci, h)
            c.fill = _fill(AZ); c.font = _font(bold=True)
            c.alignment = _align("left" if ci in (1, 3) else "center")
            c.border = _border()
        ws_r.row_dimensions[2].height = 16
        prev_fac = None
        for ri, (_, row) in enumerate(pivot_r.iterrows(), 3):
            fac = row["FACULTAD NOMPROPIO"]; siga = row["CODIGO SIGA"]; carr = row["NOMBRE CARRERA FINAL"]
            c1 = ws_r.cell(ri, 1, fac if fac != prev_fac else "")
            c1.fill = _fill(AZ_XCL); c1.font = Font(bold=True, color=AZ, size=8, name="Calibri")
            c1.alignment = _align("left"); c1.border = _border(); prev_fac = fac
            ws_r.cell(ri, 2, int(siga) if pd.notna(siga) else "").fill = _fill(GRIS_F)
            ws_r.cell(ri, 2).font = Font(bold=True, color=AZ_MED, size=8, name="Calibri")
            ws_r.cell(ri, 2).alignment = _align("center"); ws_r.cell(ri, 2).border = _border()
            ws_r.cell(ri, 3, carr).fill = _fill(GRIS_C)
            ws_r.cell(ri, 3).font = Font(color="1E2D45", size=8, name="Calibri")
            ws_r.cell(ri, 3).alignment = _align("left"); ws_r.cell(ri, 3).border = _border()
            ci_off = 4
            for col_c in cols_coh:
                v = row.get(col_c)
                c = ws_r.cell(ri, ci_off)
                if pd.isna(v):
                    c.value = "—"; c.font = Font(color="CBD5E1", size=8, name="Calibri")
                else:
                    pct = v * 100
                    c.value = f"{pct:.0f}%"
                    col_txt = VERDE if pct >= 80 else ("7A5C00" if pct >= 60 else ROJO)
                    bg_hex  = "E8F5E9" if pct >= 80 else ("FFF8E1" if pct >= 60 else "FFEBEE")
                    c.fill = _fill(bg_hex); c.font = Font(bold=True, color=col_txt, size=8, name="Calibri")
                c.alignment = _align("center"); c.border = _border()
                ci_off += 1
            if len(cols_coh) >= 2:
                va = row.get(cols_coh[-2]); vu = row.get(cols_coh[-1])
                c = ws_r.cell(ri, ci_off)
                if pd.notna(va) and pd.notna(vu):
                    d = (vu - va) * 100
                    sign = "+" if d > 0 else ""
                    c.value = f"{sign}{d:.0f}%"
                    c.font = Font(bold=True, color=(VERDE if d > 0 else ROJO), size=8, name="Calibri")
                else:
                    c.value = "—"; c.font = Font(color="CBD5E1", size=8, name="Calibri")
                c.alignment = _align("center"); c.border = _border()
                ci_off += 1
            vals_p = [row.get(c) for c in last3 if pd.notna(row.get(c))]
            c = ws_r.cell(ri, ci_off)
            if vals_p:
                prom = sum(vals_p) / len(vals_p) * 100
                bg_hex = "C8E6C9" if prom >= 80 else ("FFF9C4" if prom >= 60 else "FFCDD2")
                col_txt = VERDE if prom >= 80 else ("7A5C00" if prom >= 60 else ROJO)
                c.value = f"{prom:.0f}%"; c.fill = _fill(bg_hex)
                c.font = Font(bold=True, color=col_txt, size=8, name="Calibri")
            else:
                c.value = "—"; c.font = Font(color="CBD5E1", size=8, name="Calibri")
            c.alignment = _align("center"); c.border = _border()
            ws_r.row_dimensions[ri].height = 14
        ws_r.column_dimensions["A"].width = 22
        ws_r.column_dimensions["B"].width = 10
        ws_r.column_dimensions["C"].width = 34
        for ci in range(4, 4 + len(cols_coh) + 2):
            ws_r.column_dimensions[get_column_letter(ci)].width = 10
        ws_r.freeze_panes = "D3"

    if st.session_state.get("export_oferta") is not None:
        df_of = st.session_state["export_oferta"]
        ws_of = wb.create_sheet("Oferta Académica")
        # Título
        ws_of.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws_of.cell(1, 1, "OFERTA ACADÉMICA UAH").fill = _fill(AZ)
        ws_of.cell(1, 1).font = _font(bold=True, size=10)
        ws_of.cell(1, 1).alignment = _align("center")
        ws_of.row_dimensions[1].height = 18
        # Headers
        hdrs = ["AÑO", "CÓD. DEMRE", "CARRERA", "VAC. 1ER SEM.", "VAC. 2DO SEM.", "VIGENCIA"]
        for ci, h in enumerate(hdrs, 1):
            c = ws_of.cell(2, ci, h)
            c.fill = _fill(AZ); c.font = _font(bold=True)
            c.alignment = _align("left" if ci == 3 else "center")
            c.border = _border()
        ws_of.row_dimensions[2].height = 16
        # Datos
        for ri, (_, row) in enumerate(df_of.iterrows(), 3):
            vals = [int(row["AÑO"]), int(row["Demre"]) if row["Demre"] else "—",
                    row["Nombre Carrera"],
                    int(row["Vacantes Semestre Uno"]), int(row["Vacantes Semestre Dos"]),
                    row["Vigencia"]]
            for ci, val in enumerate(vals, 1):
                c = ws_of.cell(ri, ci, val)
                c.border = _border()
                c.alignment = _align("left" if ci == 3 else "center")
                if ci == 3:
                    c.fill = _fill(GRIS_C); c.font = Font(color="1E2D45", size=8, name="Calibri")
                elif ci in (4, 5):
                    v = int(val) if isinstance(val, int) else 0
                    color = AZ if v > 0 else "CBD5E1"
                    c.font = Font(bold=(v > 0), color=color, size=8, name="Calibri")
                else:
                    c.font = Font(color="2C3A50", size=8, name="Calibri")
            ws_of.row_dimensions[ri].height = 14
        ws_of.column_dimensions["A"].width = 8
        ws_of.column_dimensions["B"].width = 12
        ws_of.column_dimensions["C"].width = 42
        ws_of.column_dimensions["D"].width = 14
        ws_of.column_dimensions["E"].width = 14
        ws_of.column_dimensions["F"].width = 28
        ws_of.freeze_panes = "A3"

    wb.save(buf_all)
    n_ready = len(sheets_ready)
    n_total = len(all_sheets)
    nota = "" if n_ready == n_total else f" ({n_ready}/{n_total} tablas — visita las pestañas restantes para completar)"
    st.markdown(
        dl_link(buf_all.getvalue(), "UAH_en_cifras_completo.xlsx",
                f"⬇  Exportar informe completo UAH en Cifras (Excel){nota}",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        unsafe_allow_html=True
    )

# ══════════════════════════════════════════════════════════════════════════════
# PESTAÑA PRINCIPAL: OFERTA ACADÉMICA
# ══════════════════════════════════════════════════════════════════════════════
with main_oferta:

    st.markdown('<div class="upload-panel"><h4>📂 Archivo de Oferta Académica</h4><p style="font-size:.8rem;color:#6B7A8D;margin:-6px 0 10px 0">📎 Ejemplo: <code>Oferta_Academica_2010_al_2026_SIES_WEB.xlsx</code> — descarga desde <a href="https://www.mifuturo.cl" target="_blank">mifuturo.cl</a></p>', unsafe_allow_html=True)
    col_up_o, col_sh_o = st.columns([3, 2])
    with col_up_o:
        uploaded_oferta = st.file_uploader("Sube el Excel de Oferta Académica SIES", type=["xlsx"],
                                           label_visibility="visible", key="up_oferta")
    with col_sh_o:
        if not uploaded_oferta:
            st.info("El servidor SIES bloquea descargas directas — debes subir el archivo manualmente.")
    st.markdown('</div>', unsafe_allow_html=True)

    if not uploaded_oferta:
        st.info("⬆️  Sube el archivo Excel de Oferta Académica SIES para comenzar.")
    else:
        file_bytes_oferta = uploaded_oferta.read()

        @st.cache_data(show_spinner=False)
        def procesar_oferta(file_bytes: bytes) -> pd.DataFrame:
            df = pd.read_excel(io.BytesIO(file_bytes))
            df = df[df["Nombre IES"].str.contains("HURTADO", case=False, na=False)].copy()
            df = df[df["Nivel Global"] == "Pregrado"].copy()
            df["AÑO"] = df["Año"].str.extract(r"OFE_(\d{4})").astype(int)
            df["Demre"] = pd.to_numeric(df["Demre"], errors="coerce").fillna(0).astype(int)
            return df[["AÑO", "Demre", "Nombre Carrera", "Vacantes Semestre Uno",
                        "Vacantes Semestre Dos", "Vigencia"]].sort_values(
                ["AÑO", "Nombre Carrera"]).reset_index(drop=True)

        try:
            with st.spinner("Procesando oferta académica…"):
                df_oferta = procesar_oferta(file_bytes_oferta)
        except Exception as e:
            st.error(f"No se pudo procesar.\n\n`{e}`"); st.stop()

        years_of = sorted(df_oferta["AÑO"].unique())

        of1, of2, of3 = st.columns([1, 2, 2])
        with of1:
            yr_sel_of = st.selectbox("Año", [int(y) for y in years_of],
                                     index=len(years_of)-1, key="yr_oferta")
        with of2:
            busq_of = st.text_input("🔍  Buscar carrera", placeholder="Ej: Derecho…", key="busq_oferta")
        with of3:
            solo_vigentes = st.toggle("Solo con vacantes", value=True, key="tog_oferta")

        df_of = df_oferta[df_oferta["AÑO"] == yr_sel_of].copy()
        if busq_of:
            df_of = df_of[df_of["Nombre Carrera"].str.contains(busq_of, case=False, na=False)]
        if solo_vigentes:
            df_of = df_of[(df_of["Vacantes Semestre Uno"] > 0) | (df_of["Vacantes Semestre Dos"] > 0)]

        total_vac1 = int(df_of["Vacantes Semestre Uno"].sum())
        total_vac2 = int(df_of["Vacantes Semestre Dos"].sum())

        st.markdown(f"""
        <div class="metric-strip">
          <div class="metric-card"><div class="mc-label">Carreras</div>
            <div class="mc-value">{len(df_of)}</div></div>
          <div class="metric-card accent"><div class="mc-label">Vacantes 1er Sem.</div>
            <div class="mc-value">{total_vac1:,}</div></div>
          <div class="metric-card"><div class="mc-label">Vacantes 2do Sem.</div>
            <div class="mc-value">{total_vac2:,}</div></div>
          <div class="metric-card"><div class="mc-label">Total Vacantes</div>
            <div class="mc-value">{total_vac1+total_vac2:,}</div></div>
          <div class="metric-card"><div class="mc-label">Años disponibles</div>
            <div class="mc-value">{int(years_of[0])}–{int(years_of[-1])}</div></div>
        </div>""", unsafe_allow_html=True)

        def build_oferta(df, tv1, tv2):
            h = ['<table class="t1"><thead><tr>']
            h.append('<th class="c-siga">CÓD. DEMRE</th>')
            h.append('<th class="c-carrera">CARRERA</th>')
            h.append('<th>Vacantes 1er Sem.</th>')
            h.append('<th>Vacantes 2do Sem.</th>')
            h.append('<th>Total</th>')
            h.append('<th style="min-width:180px">Vigencia</th>')
            h.append('</tr></thead><tbody>')
            for _, row in df.iterrows():
                demre = int(row["Demre"]) if row["Demre"] else "—"
                carr  = row["Nombre Carrera"]
                v1    = int(row["Vacantes Semestre Uno"])
                v2    = int(row["Vacantes Semestre Dos"])
                tot   = v1 + v2
                vig   = str(row["Vigencia"])
                v1_s  = f"{v1:,}" if v1 > 0 else "—"
                v2_s  = f"{v2:,}" if v2 > 0 else "—"
                tot_s = f"{tot:,}" if tot > 0 else "—"
                vig_col = "color:#1A7A3C;font-weight:600" if "con estudiantes nuevos" in vig else "color:#999"
                h.append('<tr>')
                h.append(f'<td class="c-siga">{demre}</td>')
                h.append(f'<td class="c-carrera">{carr}</td>')
                h.append(f'<td class="c-num">{v1_s}</td>')
                h.append(f'<td class="c-num">{v2_s}</td>')
                h.append(f'<td class="c-num" style="font-weight:700">{tot_s}</td>')
                h.append(f'<td style="font-size:.75rem;{vig_col}">{vig}</td>')
                h.append('</tr>')
            h.append('<tr class="tr-tot"><td class="c-fac" colspan="2">TOTAL UAH</td>')
            h.append(f'<td>{tv1:,}</td><td>{tv2:,}</td><td>{tv1+tv2:,}</td><td></td></tr>')
            h.append('</tbody></table>')
            return "".join(h)

        st.markdown(f"""
        <div class="table-wrapper">
          <div class="table-title-bar">
            OFERTA ACADÉMICA UAH — {yr_sel_of}
            &nbsp;·&nbsp; {len(df_of)} carreras
          </div>
          <div class="table-scroll">{build_oferta(df_of, total_vac1, total_vac2)}</div>
        </div>""", unsafe_allow_html=True)

        df_dl_of = df_of.copy()
        df_dl_of["Total Vacantes"] = df_dl_of["Vacantes Semestre Uno"] + df_dl_of["Vacantes Semestre Dos"]
        df_dl_of.columns = ["AÑO", "CÓD. DEMRE", "CARRERA", "VAC. 1ER SEM.", "VAC. 2DO SEM.", "VIGENCIA", "TOTAL VACANTES"]
        st.session_state["export_oferta"] = df_of
        buf_of = io.BytesIO()
        with pd.ExcelWriter(buf_of, engine="openpyxl") as w:
            df_dl_of.to_excel(w, index=False, sheet_name="Oferta Académica UAH")
        st.markdown(
            dl_link(buf_of.getvalue(), "UAH_oferta_academica.xlsx",
                    "⬇  Descargar Excel",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            + dl_link(df_dl_of.to_csv(index=False).encode("utf-8-sig"),
                      "UAH_oferta_academica.csv", "⬇  Descargar CSV", "text/csv"),
            unsafe_allow_html=True
        )